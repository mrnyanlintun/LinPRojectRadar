#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

import networkx as nx
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import linprog

SEED = 20260811
PROGRAMME_VERSION = "OG-SYNTH-0.2"
GENERATOR_VERSION = "build_opus_synthetic_programme_v0_2.py@0.2"
DATA_ORIGIN = "SYNTHETIC_RESEARCH_FIXTURE"
GENERATED_AT = "2026-08-11T00:00:00Z"
FIXED_DT = datetime(2026, 8, 11, 0, 0, 0)
PROVENANCE_COLUMNS = [
    "data_origin",
    "programme_version",
    "package_version",
    "generator_version",
    "random_seed",
    "not_for_empirical_validation",
    "record_hash",
]
CATEGORY_NAMES = {
    1: "Quantitative EVM / Cost & Performance Forecasting",
    2: "Schedule Analytics",
    3: "Cost Risk",
    4: "Document & Risk Signals",
    5: "System Dynamics & Complexity",
    6: "Signal Synthesis",
    7: "Evidence Combination / Epistemic Uncertainty",
    8: "Governance & Compliance",
    9: "Data Integrity & Information Quality",
    10: "Decision Optimization",
}
CODE_ID_SPECIAL = {"8.8": "A6.3"}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def json_default(value: Any) -> Any:
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    raise TypeError(type(value))


def row_hash(row: pd.Series) -> str:
    payload = {
        key: (None if pd.isna(value) else value)
        for key, value in row.items()
        if key != "record_hash"
    }
    raw = json.dumps(payload, sort_keys=True, default=json_default, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def package_version_for(path: Path, root: Path) -> str:
    rel = path.relative_to(root)
    parts = rel.parts
    if parts and parts[0] == "package_A_project_structures":
        return "A-0.2"
    if parts[:2] == ("package_B_reference_training_decisions", "B1_reference_population"):
        return "B1-0.2"
    if parts[:2] == ("package_B_reference_training_decisions", "B2_expert_epistemic"):
        return "B2-0.2"
    if parts[:2] == ("package_B_reference_training_decisions", "B3_decision_optimization"):
        return "B3-0.2"
    if parts and parts[0] == "package_C_optional_activation_lab":
        return "C-0.2"
    return "ROOT-0.2"


def strip_provenance(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop(columns=[column for column in PROVENANCE_COLUMNS if column in df.columns])


def add_provenance(df: pd.DataFrame, package_version: str) -> pd.DataFrame:
    out = strip_provenance(df.copy())
    out["data_origin"] = DATA_ORIGIN
    out["programme_version"] = PROGRAMME_VERSION
    out["package_version"] = package_version
    out["generator_version"] = GENERATOR_VERSION
    out["random_seed"] = SEED
    out["not_for_empirical_validation"] = True
    out["record_hash"] = out.apply(row_hash, axis=1)
    return out


def write_csv(df: pd.DataFrame, path: Path, root: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    add_provenance(df, package_version_for(path, root)).to_csv(path, index=False)


def write_json(value: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, indent=2, sort_keys=True, default=json_default), encoding="utf-8"
    )


def code_module_id_for(literature_module_id: str) -> str:
    if literature_module_id in CODE_ID_SPECIAL:
        return CODE_ID_SPECIAL[literature_module_id]
    category, index = literature_module_id.split(".", 1)
    category_number = int(category)
    if 1 <= category_number <= 5:
        return f"A{category_number}.{index}"
    if category_number == 6:
        return f"B1.{index}"
    if category_number == 7:
        return f"B2.{index}"
    if category_number == 10:
        return f"B4.{index}"
    return literature_module_id


def safe_extract(zip_path: Path, target: Path) -> Path:
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            name = info.filename
            path = Path(name)
            if path.is_absolute() or ".." in path.parts:
                raise RuntimeError(f"unsafe archive path: {name}")
        archive.extractall(target)
    roots = [path for path in target.iterdir() if path.is_dir()]
    if len(roots) != 1:
        raise RuntimeError(f"expected one archive root, got {roots}")
    return roots[0]


def refresh_csv_provenance(root: Path) -> None:
    excluded = {"MANIFEST.csv", "PACKAGE_MANIFEST.csv", "data_dictionary.csv"}
    for path in sorted(root.rglob("*.csv")):
        if path.name in excluded:
            continue
        header = pd.read_csv(path, nrows=0)
        id_columns = {
            column: str
            for column in ["module_id", "literature_module_id", "code_module_id"]
            if column in header.columns
        }
        df = pd.read_csv(path, dtype=id_columns or None)
        add_provenance(df, package_version_for(path, root)).to_csv(path, index=False)


def update_json_versions(value: Any) -> Any:
    if isinstance(value, dict):
        out = {key: update_json_versions(item) for key, item in value.items()}
        if "programme_version" in out:
            out["programme_version"] = PROGRAMME_VERSION
        if "data_origin" in out:
            out["data_origin"] = DATA_ORIGIN
        if "not_for_empirical_validation" in out:
            out["not_for_empirical_validation"] = True
        return out
    if isinstance(value, list):
        return [update_json_versions(item) for item in value]
    if value == "OG-SYNTH-0.1":
        return PROGRAMME_VERSION
    return value


def refresh_json_versions(root: Path) -> None:
    for path in sorted(root.rglob("*.json")):
        try:
            value = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        write_json(update_json_versions(value), path)


def build_ccpm(root: Path) -> None:
    package = root / "package_A_project_structures"
    activities = pd.read_csv(package / "schedule_activities.csv")
    dependencies = pd.read_csv(package / "schedule_dependencies.csv")
    status = pd.read_csv(package / "schedule_status.csv")
    periods = pd.read_csv(package / "reporting_periods.csv")
    schedule_gt = pd.read_csv(package / "schedule_ground_truth.csv")

    chain_rows: list[dict[str, Any]] = []
    chain_activity_rows: list[dict[str, Any]] = []
    sizing_rows: list[dict[str, Any]] = []
    buffer_rows: list[dict[str, Any]] = []
    buffer_gt_rows: list[dict[str, Any]] = []
    activity_chain_map: dict[tuple[str, str], tuple[str, str]] = {}

    for project_id, activity_group in activities.groupby("project_id", sort=True):
        activity_group = activity_group.sort_values("activity_id").copy()
        dependency_group = dependencies[dependencies.project_id == project_id]
        graph = nx.DiGraph()
        graph.add_nodes_from(activity_group.activity_id)
        graph.add_edges_from(
            zip(dependency_group.predecessor_activity_id, dependency_group.successor_activity_id)
        )
        topological = list(nx.topological_sort(graph))
        critical = [
            activity_id
            for activity_id in topological
            if bool(
                activity_group[activity_group.activity_id == activity_id]
                .critical_chain_flag.iloc[0]
            )
        ]
        noncritical = [activity_id for activity_id in topological if activity_id not in critical]
        feeding_one = noncritical[::2]
        feeding_two = noncritical[1::2]
        critical_target_mid = critical[min(len(critical) - 1, max(0, len(critical) // 2))] if critical else topological[-1]
        chain_definitions = [
            (f"{project_id}-CC", "PROJECT", critical, critical[-1] if critical else topological[-1]),
            (f"{project_id}-FC1", "FEEDING", feeding_one, critical_target_mid),
            (f"{project_id}-FC2", "FEEDING", feeding_two, critical[-1] if critical else topological[-1]),
        ]
        chain_sizes: dict[str, tuple[float, float]] = {}
        for chain_id, chain_type, members, feeds_into in chain_definitions:
            if not members:
                continue
            variances: list[float] = []
            for sequence, activity_id in enumerate(members, 1):
                row = activity_group[activity_group.activity_id == activity_id].iloc[0]
                sigma = (
                    float(row.pessimistic_duration_days)
                    - float(row.optimistic_duration_days)
                ) / 6.0
                variance = sigma**2
                variances.append(variance)
                activity_chain_map[(project_id, activity_id)] = (chain_id, chain_type)
                chain_activity_rows.append(
                    {
                        "project_id": project_id,
                        "chain_id": chain_id,
                        "chain_type": chain_type,
                        "activity_id": activity_id,
                        "chain_sequence": sequence,
                        "feeds_into_activity_id": feeds_into if chain_type == "FEEDING" else "",
                    }
                )
                sizing_rows.append(
                    {
                        "project_id": project_id,
                        "chain_id": chain_id,
                        "activity_id": activity_id,
                        "optimistic_duration_days": float(row.optimistic_duration_days),
                        "most_likely_duration_days": float(row.most_likely_duration_days),
                        "pessimistic_duration_days": float(row.pessimistic_duration_days),
                        "pert_sigma_days": round(sigma, 8),
                        "variance_days2": round(variance, 8),
                        "buffer_sizing_method": "RSS_PERT_VARIANCE",
                        "sizing_z": 1.645,
                    }
                )
            variance_sum = float(sum(variances))
            buffer_days = max(1.0, 1.645 * math.sqrt(variance_sum))
            chain_sizes[chain_id] = (buffer_days, variance_sum)
            chain_rows.append(
                {
                    "project_id": project_id,
                    "chain_id": chain_id,
                    "chain_type": chain_type,
                    "feeds_into_activity_id": feeds_into if chain_type == "FEEDING" else "",
                    "activity_count": len(members),
                    "variance_sum_days2": round(variance_sum, 8),
                    "original_buffer_days": round(buffer_days, 8),
                    "buffer_sizing_method": "RSS_PERT_VARIANCE",
                    "sizing_quantile": "P95_ONE_SIDED",
                    "sizing_z": 1.645,
                }
            )

        baseline_finish = float(
            schedule_gt[schedule_gt.project_id == project_id].baseline_finish_day.iloc[0]
        )
        for _, period in periods[periods.project_id == project_id].sort_values("period_number").iterrows():
            period_id = period.period_id
            period_status = status[
                (status.project_id == project_id) & (status.period_id == period_id)
            ]
            current_finish = float(
                schedule_gt[
                    (schedule_gt.project_id == project_id)
                    & (schedule_gt.period_id == period_id)
                ].current_forecast_finish_day.iloc[0]
            )
            delay_component = max(0.0, current_finish - baseline_finish)
            project_consumed = 0.0
            project_original = 0.0
            project_progress = 0.0
            for chain_id, chain_type, members, _ in chain_definitions:
                if not members or chain_id not in chain_sizes:
                    continue
                original, variance_sum = chain_sizes[chain_id]
                progress = float(
                    period_status[period_status.activity_id.isin(members)]
                    .progress_fraction.mean()
                )
                burn_factor = 0.62 if chain_type == "PROJECT" else 0.48
                delay_factor = 0.18 if chain_type == "PROJECT" else 0.08
                consumed = min(
                    original,
                    original * progress * burn_factor + delay_component * delay_factor,
                )
                remaining = max(0.0, original - consumed)
                buffer_rows.append(
                    {
                        "project_id": project_id,
                        "period_id": period_id,
                        "buffer_id": f"{chain_id}-BUFFER",
                        "buffer_type": "PROJECT" if chain_type == "PROJECT" else "FEEDING",
                        "chain_id": chain_id,
                        "original_buffer_days": round(original, 8),
                        "remaining_buffer_days": round(remaining, 8),
                        "chain_progress_fraction": round(progress, 8),
                        "buffer_sizing_method": "RSS_PERT_VARIANCE",
                        "variance_sum_days2": round(variance_sum, 8),
                        "sizing_z": 1.645,
                    }
                )
                if chain_type == "PROJECT":
                    project_consumed = consumed
                    project_original = original
                    project_progress = progress
            buffer_gt_rows.append(
                {
                    "project_id": project_id,
                    "period_id": period_id,
                    "project_buffer_consumed_days": round(project_consumed, 8),
                    "project_buffer_consumption_ratio": round(
                        project_consumed / project_original if project_original else 0.0,
                        8,
                    ),
                    "buffer_to_progress_ratio": round(
                        (project_consumed / project_original) / project_progress
                        if project_original and project_progress > 0
                        else 0.0,
                        8,
                    ),
                    "buffer_exhausted": bool(
                        project_original and project_consumed >= project_original - 1e-9
                    ),
                }
            )

    activities = strip_provenance(activities)
    activities["ccpm_chain_id"] = [
        activity_chain_map.get((row.project_id, row.activity_id), ("", ""))[0]
        for row in activities.itertuples()
    ]
    activities["ccpm_chain_type"] = [
        activity_chain_map.get((row.project_id, row.activity_id), ("", ""))[1]
        for row in activities.itertuples()
    ]
    write_csv(activities, package / "schedule_activities.csv", root)
    write_csv(pd.DataFrame(chain_rows), package / "ccpm_chains.csv", root)
    write_csv(pd.DataFrame(chain_activity_rows), package / "ccpm_chain_activities.csv", root)
    write_csv(pd.DataFrame(sizing_rows), package / "ccpm_buffer_sizing_inputs.csv", root)
    write_csv(pd.DataFrame(buffer_rows), package / "ccpm_buffers.csv", root)
    write_csv(pd.DataFrame(buffer_gt_rows), package / "ccpm_ground_truth.csv", root)


def build_abm_rules(root: Path) -> None:
    package = root / "package_A_project_structures"
    rules = pd.DataFrame(
        [
            {
                "decision_rule_id": "RESTOCK_AND_PRIORITIZE_CRITICAL",
                "rule_order": 1,
                "rule_branch": "DISRUPTION_CAPACITY_REDUCTION",
                "condition_json": json.dumps({"state": "DISRUPTED"}, sort_keys=True),
                "action_json": json.dumps(
                    {
                        "capacity_multiplier": 0.45,
                        "priority": "CRITICAL_FIRST",
                        "restock_fraction_of_base_capacity": 0.75,
                    },
                    sort_keys=True,
                ),
                "next_state": "DISRUPTED",
                "rule_version": "ABM-RULES-0.2",
            },
            {
                "decision_rule_id": "RESTOCK_AND_PRIORITIZE_CRITICAL",
                "rule_order": 2,
                "rule_branch": "LOW_INVENTORY_RESTOCK",
                "condition_json": json.dumps(
                    {"inventory_below_base_capacity": True}, sort_keys=True
                ),
                "action_json": json.dumps(
                    {
                        "priority": "CRITICAL_FIRST",
                        "restock_fraction_of_base_capacity": 0.75,
                    },
                    sort_keys=True,
                ),
                "next_state": "NORMAL",
                "rule_version": "ABM-RULES-0.2",
            },
            {
                "decision_rule_id": "RESTOCK_AND_PRIORITIZE_CRITICAL",
                "rule_order": 3,
                "rule_branch": "NORMAL_FULFILMENT",
                "condition_json": json.dumps({"default": True}, sort_keys=True),
                "action_json": json.dumps(
                    {
                        "capacity_multiplier": 1.0,
                        "priority": "FCFS",
                        "restock_fraction_of_base_capacity": 0.75,
                    },
                    sort_keys=True,
                ),
                "next_state": "NORMAL",
                "rule_version": "ABM-RULES-0.2",
            },
        ]
    )
    write_csv(rules, package / "agent_decision_rules.csv", root)
    history = pd.read_csv(package / "agent_state_history.csv")
    history = strip_provenance(history)
    history["decision_rule_id"] = "RESTOCK_AND_PRIORITIZE_CRITICAL"
    history["rule_branch"] = np.where(
        history.state == "DISRUPTED",
        "DISRUPTION_CAPACITY_REDUCTION",
        np.where(
            history.inventory_end_units < history.effective_capacity_units,
            "LOW_INVENTORY_RESTOCK",
            "NORMAL_FULFILMENT",
        ),
    )
    write_csv(history, package / "agent_state_history.csv", root)
    ground_truth = (
        history.groupby(["project_id", "decision_rule_id", "rule_branch"], as_index=False)
        .size()
        .rename(columns={"size": "application_count"})
    )
    write_csv(ground_truth, package / "abm_rule_ground_truth.csv", root)


def build_ncr_corpus(root: Path) -> None:
    package = root / "package_A_project_structures"
    projects = pd.read_csv(package / "projects.csv")
    periods = pd.read_csv(package / "reporting_periods.csv")
    rng = np.random.default_rng(SEED + 301)
    audits: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []
    ground_truth: list[dict[str, Any]] = []

    for project_id in sorted(projects.project_id):
        project_periods = periods[periods.project_id == project_id].sort_values("period_number")
        project_events: list[dict[str, Any]] = []
        cumulative_inspections = 0
        sequence = 1
        for _, period in project_periods.iterrows():
            period_number = int(period.period_number)
            cutoff = pd.Timestamp(period.period_end)
            inspections = 18 + period_number * 3 + int(rng.integers(0, 5))
            cumulative_inspections += inspections
            new_count = int(max(0, rng.poisson(1.2 + 0.25 * period_number)))
            audit_id = f"{project_id}-{period.period_id}-QA"
            severities: list[str] = []
            for _ in range(new_count):
                severity = str(
                    rng.choice(["MINOR", "MAJOR", "CRITICAL"], p=[0.62, 0.30, 0.08])
                )
                severities.append(severity)
                issue_date = cutoff - pd.Timedelta(days=int(rng.integers(0, 18)))
                closure_lag = int(rng.choice([0, 1, 2, 99], p=[0.36, 0.32, 0.20, 0.12]))
                close_date: pd.Timestamp | None = None
                if closure_lag != 99 and period_number + closure_lag <= 6:
                    close_cutoff = pd.Timestamp(
                        project_periods[
                            project_periods.period_number == period_number + closure_lag
                        ].period_end.iloc[0]
                    )
                    close_date = min(
                        close_cutoff,
                        issue_date + pd.Timedelta(days=10 + 8 * closure_lag),
                    )
                event = {
                    "project_id": project_id,
                    "ncr_id": f"{project_id}-NCR-{sequence:03d}",
                    "source_audit_id": audit_id,
                    "issue_period_id": period.period_id,
                    "issue_date": issue_date.date().isoformat(),
                    "due_date": (
                        issue_date
                        + pd.Timedelta(days=21 if severity == "MINOR" else 14)
                    ).date().isoformat(),
                    "close_date": close_date.date().isoformat() if close_date else "",
                    "severity": severity,
                    "description_code": str(
                        rng.choice(
                            [
                                "WORKMANSHIP",
                                "MATERIAL",
                                "DOCUMENTATION",
                                "TEST_FAILURE",
                                "INSTALLATION",
                            ]
                        )
                    ),
                }
                sequence += 1
                project_events.append(event)
                events.append(event)
            audits.append(
                {
                    "project_id": project_id,
                    "period_id": period.period_id,
                    "audit_id": audit_id,
                    "audit_date": cutoff.date().isoformat(),
                    "inspections_completed": inspections,
                    "work_packages_audited": max(1, inspections // 3),
                    "total_findings": new_count + int(rng.integers(0, 3)),
                    "critical_findings": sum(
                        1 for severity in severities if severity == "CRITICAL"
                    ),
                    "new_ncr_count": new_count,
                }
            )
            issued = [
                event
                for event in project_events
                if pd.Timestamp(event["issue_date"]) <= cutoff
            ]
            closed = [
                event
                for event in issued
                if event["close_date"]
                and pd.Timestamp(event["close_date"]) <= cutoff
            ]
            open_items = [
                event
                for event in issued
                if not event["close_date"]
                or pd.Timestamp(event["close_date"]) > cutoff
            ]
            overdue = [
                event
                for event in open_items
                if pd.Timestamp(event["due_date"]) < cutoff
            ]
            ages = [
                (cutoff - pd.Timestamp(event["issue_date"])).days
                for event in open_items
            ]
            ground_truth.append(
                {
                    "project_id": project_id,
                    "period_id": period.period_id,
                    "ncr_issued_to_date": len(issued),
                    "ncr_closed_to_date": len(closed),
                    "ncr_open_at_cutoff": len(open_items),
                    "ncr_overdue_at_cutoff": len(overdue),
                    "cumulative_inspections": cumulative_inspections,
                    "ncr_incidence_per_100_inspections": round(
                        100.0 * len(issued) / cumulative_inspections, 8
                    ),
                    "closure_ratio": round(len(closed) / len(issued), 8)
                    if issued
                    else 0.0,
                    "open_ratio": round(len(open_items) / len(issued), 8)
                    if issued
                    else 0.0,
                    "overdue_open_ratio": round(len(overdue) / len(open_items), 8)
                    if open_items
                    else 0.0,
                    "mean_open_age_days": round(float(np.mean(ages)), 8)
                    if ages
                    else 0.0,
                }
            )
    write_csv(pd.DataFrame(audits), package / "quality_audits.csv", root)
    write_csv(pd.DataFrame(events), package / "ncr_events.csv", root)
    write_csv(pd.DataFrame(ground_truth), package / "ncr_ground_truth.csv", root)


def build_environmental_corpus(root: Path) -> None:
    package = root / "package_A_project_structures"
    projects = pd.read_csv(package / "projects.csv")
    periods = pd.read_csv(package / "reporting_periods.csv")
    rng = np.random.default_rng(SEED + 401)
    templates = [
        ("STORMWATER", "Stormwater controls", "HIGH"),
        ("EROSION", "Erosion and sediment control", "HIGH"),
        ("WASTE", "Waste segregation and disposal", "MEDIUM"),
        ("DUST", "Dust and air-quality control", "MEDIUM"),
        ("NOISE", "Noise and work-hour control", "LOW"),
        ("SPILL", "Spill prevention and response", "CRITICAL"),
        ("DEWATERING", "Dewatering discharge control", "HIGH"),
        ("MATERIAL_STORAGE", "Hazardous-material storage", "CRITICAL"),
    ]
    requirements: list[dict[str, Any]] = []
    assessments: list[dict[str, Any]] = []
    violations: list[dict[str, Any]] = []
    ground_truth: list[dict[str, Any]] = []

    for project_id in sorted(projects.project_id):
        project_periods = periods[periods.project_id == project_id].sort_values("period_number")
        local_requirements: list[dict[str, Any]] = []
        for index, (code, name, severity) in enumerate(templates, 1):
            requirement_id = f"{project_id}-ENV-{index:02d}"
            applicable = not (
                code == "DEWATERING" and project_id in {"PRJ-DCT", "PRJ-AIR"}
            )
            row = {
                "project_id": project_id,
                "requirement_id": requirement_id,
                "requirement_code": code,
                "requirement_name": name,
                "permit_or_rule_source": f"SYNTHETIC-PERMIT-{project_id}",
                "jurisdiction": "SYNTHETIC-US",
                "severity_if_breached": severity,
                "applicable": applicable,
                "effective_date": project_periods.period_start.min(),
            }
            local_requirements.append(row)
            requirements.append(row)

        project_violations: list[dict[str, Any]] = []
        for _, period in project_periods.iterrows():
            period_number = int(period.period_number)
            assessed_count = 0
            compliant_count = 0
            noncompliant_count = 0
            unassessed_count = 0
            severe_count = 0
            for index, requirement in enumerate(local_requirements, 1):
                if not requirement["applicable"]:
                    result = "NOT_APPLICABLE"
                else:
                    draw = float(rng.random())
                    if draw < 0.07 + 0.01 * period_number:
                        result = "NONCOMPLIANT"
                    elif draw < 0.13:
                        result = "NOT_ASSESSED"
                    else:
                        result = "COMPLIANT"
                assessment_id = (
                    f"{requirement['requirement_id']}-{period.period_id}"
                )
                assessments.append(
                    {
                        "project_id": project_id,
                        "period_id": period.period_id,
                        "assessment_id": assessment_id,
                        "requirement_id": requirement["requirement_id"],
                        "assessment_date": period.period_end,
                        "result": result,
                        "evidence_code": f"ENV-EVID-{project_id}-{period.period_id}-{index:02d}",
                        "reviewer_role": "ENVIRONMENTAL_COMPLIANCE_REVIEWER",
                    }
                )
                if result in {"COMPLIANT", "NONCOMPLIANT"}:
                    assessed_count += 1
                if result == "COMPLIANT":
                    compliant_count += 1
                elif result == "NONCOMPLIANT":
                    noncompliant_count += 1
                    severity = requirement["severity_if_breached"]
                    severe_count += int(severity in {"HIGH", "CRITICAL"})
                    assessment_date = pd.Timestamp(period.period_end)
                    due_date = assessment_date + pd.Timedelta(
                        days=14 if severity == "CRITICAL" else 30
                    )
                    closed = bool(rng.random() < 0.55)
                    close_date = (
                        assessment_date + pd.Timedelta(days=int(rng.integers(3, 35)))
                        if closed
                        else None
                    )
                    violation = {
                        "project_id": project_id,
                        "period_id": period.period_id,
                        "violation_id": f"{assessment_id}-V",
                        "assessment_id": assessment_id,
                        "requirement_id": requirement["requirement_id"],
                        "severity": severity,
                        "identified_date": assessment_date.date().isoformat(),
                        "corrective_due_date": due_date.date().isoformat(),
                        "corrective_close_date": close_date.date().isoformat()
                        if close_date
                        else "",
                        "status": "CLOSED" if closed else "OPEN",
                    }
                    project_violations.append(violation)
                    violations.append(violation)
                elif result == "NOT_ASSESSED":
                    unassessed_count += 1
            cutoff = pd.Timestamp(period.period_end)
            overdue_actions = sum(
                1
                for violation in project_violations
                if violation["status"] == "OPEN"
                and pd.Timestamp(violation["corrective_due_date"]) < cutoff
            )
            applicable_count = sum(
                1 for requirement in local_requirements if requirement["applicable"]
            )
            ground_truth.append(
                {
                    "project_id": project_id,
                    "period_id": period.period_id,
                    "applicable_requirements": applicable_count,
                    "applicable_requirements_assessed": assessed_count,
                    "compliant_requirements": compliant_count,
                    "noncompliant_requirements": noncompliant_count,
                    "unassessed_requirements": unassessed_count,
                    "environmental_compliance_rate": round(
                        compliant_count / assessed_count, 8
                    )
                    if assessed_count
                    else 0.0,
                    "severe_noncompliances": severe_count,
                    "overdue_corrective_actions": overdue_actions,
                }
            )
    write_csv(pd.DataFrame(requirements), package / "environmental_requirements.csv", root)
    write_csv(pd.DataFrame(assessments), package / "environmental_assessments.csv", root)
    write_csv(pd.DataFrame(violations), package / "environmental_violations.csv", root)
    write_csv(pd.DataFrame(ground_truth), package / "environmental_ground_truth.csv", root)


def build_lp_models(root: Path) -> None:
    package = root / "package_B_reference_training_decisions" / "B3_decision_optimization"
    problems = pd.read_csv(package / "decision_problems.csv")
    ground_truth = pd.read_csv(package / "ground_truth_decisions.csv")
    models: list[dict[str, Any]] = []
    for _, problem in problems.iterrows():
        problem_id = str(problem.decision_problem_id)
        number = int(problem_id.split("-")[-1])
        stress = 0.75 + 0.05 * number
        objective = (
            np.array([1.6, 1.1, 1.4, 0.6]) * 1_000_000 * stress
            + np.array([-10, -7, -9, -6]) * 180_000 * stress
        ).tolist()
        a_ub = [
            [1.6, 1.1, 1.4, 0.6],
            [2.0, 0.0, 0.0, 0.0],
            [0.0, 0.15, 0.0, 0.0],
        ]
        b_ub = [
            float(problem.budget_cap_usd) / 1_000_000.0,
            float(problem.crew_capacity_increment),
            float(problem.max_overtime_fraction),
        ]
        result = linprog(
            np.array(objective, dtype=float),
            A_ub=np.array(a_ub, dtype=float),
            b_ub=np.array(b_ub, dtype=float),
            bounds=[(0.0, 1.0)] * 4,
            method="highs",
        )
        expected = ground_truth[
            ground_truth.decision_problem_id == problem_id
        ].iloc[0]
        if bool(result.success) != bool(expected.lp_success):
            raise RuntimeError(f"LP success mismatch for {problem_id}")
        models.append(
            {
                "decision_problem_id": problem_id,
                "model_class": "LINEAR_PROGRAM",
                "objective": {
                    "sense": "MIN",
                    "coefficients": objective,
                    "constant": 0.0,
                    "unit": "USD_EQUIVALENT",
                },
                "variables": [
                    {
                        "name": "add_crew_intensity",
                        "lower": 0.0,
                        "upper": 1.0,
                        "integrality": 0,
                    },
                    {
                        "name": "overtime_intensity",
                        "lower": 0.0,
                        "upper": 1.0,
                        "integrality": 0,
                    },
                    {
                        "name": "expedite_intensity",
                        "lower": 0.0,
                        "upper": 1.0,
                        "integrality": 0,
                    },
                    {
                        "name": "resequence_intensity",
                        "lower": 0.0,
                        "upper": 1.0,
                        "integrality": 0,
                    },
                ],
                "constraints": [
                    {
                        "name": "ACTION_COST_CAP",
                        "sense": "LE",
                        "coefficients": a_ub[0],
                        "rhs": b_ub[0],
                        "unit": "USD_MILLIONS",
                    },
                    {
                        "name": "CREW_CAPACITY",
                        "sense": "LE",
                        "coefficients": a_ub[1],
                        "rhs": b_ub[1],
                        "unit": "CREW",
                    },
                    {
                        "name": "OVERTIME_CAP",
                        "sense": "LE",
                        "coefficients": a_ub[2],
                        "rhs": b_ub[2],
                        "unit": "FRACTION",
                    },
                ],
                "solver_reference": {
                    "library": "scipy.optimize.linprog",
                    "method": "highs",
                },
                "ground_truth": {
                    "success": bool(result.success),
                    "objective_value": float(result.fun),
                    "solution": {
                        "add_crew": float(result.x[0]),
                        "overtime": float(result.x[1]),
                        "expedite": float(result.x[2]),
                        "resequence": float(result.x[3]),
                    },
                },
            }
        )
    write_json(
        {
            "schema_version": "LP-MODEL-0.2",
            "data_origin": DATA_ORIGIN,
            "programme_version": PROGRAMME_VERSION,
            "not_for_empirical_validation": True,
            "models": models,
        },
        package / "lp_models.json",
    )
    write_json(
        {
            "schema_version": "LP-MODEL-0.2",
            "required_fields": [
                "decision_problem_id",
                "model_class",
                "objective",
                "variables",
                "constraints",
                "solver_reference",
                "ground_truth",
            ],
            "coefficient_order": "matches variables array",
            "constraint_senses": ["LE", "EQ", "GE"],
        },
        package / "lp_model_schema.json",
    )


def build_module_maps(root: Path, base_map: pd.DataFrame) -> None:
    module_map = strip_provenance(base_map.copy())
    module_map["module_id"] = module_map.module_id.astype(str)
    updates = {
        "2.3": "ccpm_chains.csv|ccpm_chain_activities.csv|ccpm_buffer_sizing_inputs.csv|ccpm_buffers.csv|ccpm_ground_truth.csv",
        "5.7": "agent_decision_rules.csv|agents.csv|agent_state_history.csv|abm_ground_truth.csv|abm_rule_ground_truth.csv",
    }
    for module_id, files in updates.items():
        module_map.loc[module_map.module_id == module_id, "primary_files"] = files
    additions = pd.DataFrame(
        [
            {
                "module_id": "4.4",
                "module_name": "NCR Rate",
                "category_number": 4,
                "category_name": CATEGORY_NAMES[4],
                "synthetic_package": "A",
                "primary_files": "quality_audits.csv|ncr_events.csv|ncr_ground_truth.csv",
                "proposed_owner_action": "IMPLEMENT_OR_TEST_WITH_SYNTHETIC_FIXTURES",
            },
            {
                "module_id": "8.8",
                "module_name": "Environmental Compliance Rate",
                "category_number": 8,
                "category_name": CATEGORY_NAMES[8],
                "synthetic_package": "A",
                "primary_files": "environmental_requirements.csv|environmental_assessments.csv|environmental_violations.csv|environmental_ground_truth.csv",
                "proposed_owner_action": "IMPLEMENT_OR_TEST_WITH_SYNTHETIC_FIXTURES",
            },
        ]
    )
    for _, addition in additions.iterrows():
        if not (module_map.module_id == addition.module_id).any():
            module_map = pd.concat([module_map, pd.DataFrame([addition])], ignore_index=True)
    module_map["category_number"] = module_map.category_number.astype(int)
    module_map["category_name"] = module_map.apply(
        lambda row: CATEGORY_NAMES.get(int(row.category_number), row.category_name), axis=1
    )
    module_map.insert(
        1,
        "code_module_id",
        module_map.module_id.map(code_module_id_for),
    )
    module_map = module_map.sort_values(
        ["category_number", "module_id"], key=lambda series: series.astype(str)
    )
    write_csv(module_map, root / "module_asset_map.csv", root)

    aliases = module_map[
        [
            "module_id",
            "code_module_id",
            "module_name",
            "category_number",
            "category_name",
            "synthetic_package",
        ]
    ].rename(columns={"module_id": "literature_module_id"})
    write_csv(aliases, root / "module_id_aliases.csv", root)

    # Add code IDs to Package C activation records and reuse manifest where possible.
    package_c = root / "package_C_optional_activation_lab"
    for filename in ["activation_candidates.csv", "reuse_manifest.csv"]:
        path = package_c / filename
        if not path.exists():
            continue
        header = pd.read_csv(path, nrows=0)
        df = strip_provenance(pd.read_csv(path, dtype={"module_id": str} if "module_id" in header.columns else None))
        if "module_id" in df.columns and "code_module_id" not in df.columns:
            df.insert(1, "code_module_id", df.module_id.astype(str).map(code_module_id_for))
        write_csv(df, path, root)


def write_readmes(root: Path) -> None:
    (root / "README.md").write_text(
        """# Opus Gubernatio Synthetic Evidence and Decision Programme v0.2

This combined archive is the authoritative release. It includes Packages A, B and C, the source generator, independent validator, checksum verifier, validation report, module maps, schema catalog, manifests, and audit-resolution record.

All records are `SYNTHETIC_RESEARCH_FIXTURE` and are marked `not_for_empirical_validation = true`. These assets support implementation verification, known-answer testing, solver agreement, controlled sensitivity, abstention testing, and document-to-module traceability. They do not establish real-world calibration or field validity.

Separate Package A/B/C archives are convenience exports and contain package-local manifests and checksums.
""",
        encoding="utf-8",
    )
    package_a = root / "package_A_project_structures"
    (package_a / "README.md").write_text(
        """# Package A — Synthetic Canonical Project Structures

Six integrated synthetic projects across six reporting periods. This package supplies activity networks, stochastic durations, Line of Balance production flow, variance-sized CCPM chains and buffers, DSM dependencies, system-dynamics states, queue events, machine-readable agent rules and histories, discrete-event logs, weather/activity links, claims/dispute stages, specification conflicts, quality audits/NCR events, environmental requirement assessments, and project cost-risk structures.

DSM is intentionally located in Package A because it is a project-specific dependency structure rather than a historical reference population.
""",
        encoding="utf-8",
    )
    (root / "AUDIT_RESOLUTION_v0.2.md").write_text(
        """# Synthetic Programme v0.2 Audit Resolution

This release resolves the ingest-and-reconciliation findings reported against the separately supplied v0.1 archives.

1. The combined archive is authoritative and contains the builder, validator, validation report, programme module map, schema catalog and manifest.
2. Package-local manifests and checksums are supplied for separate-archive verification.
3. NCR Rate receives a quality-audit and event-level NCR cohort with independently recomputable ground truth.
4. Environmental Compliance Rate receives an applicable-requirement register, assessments, violations and ground truth.
5. CCPM buffers trace to explicit chains and activities and are sized from PERT variance using a declared one-sided P95 z-value rather than a flat percentage.
6. Agent decision rules are supplied as machine-readable conditions/actions and linked to every agent-state record.
7. DSM remains in Package A because it is a project-specific canonical structure. This boundary decision is explicit.
8. LP models contain numerical objectives, bounds and coefficient vectors consumable by a solver.
9. Module aliases map literature IDs such as 7.19 to code IDs such as B2.19.

All data remain synthetic research fixtures and are not empirical validation evidence.
""",
        encoding="utf-8",
    )
    (root / "CLAUDE_CODE_HANDOFF_v0.2.md").write_text(
        """# Claude Code Handoff — Synthetic Programme v0.2

Use the combined v0.2 archive as authoritative. Verify `CHECKSUMS.sha256`, rerun `generators/validate_synthetic_programme_v0_2.py`, and reconcile `module_id_aliases.csv` with the current repository registry before connecting any module.

Do not activate disabled modules, change voting, or treat synthetic fixtures as operational project evidence. Package A contains project-specific structures. Package B contains reference, training, epistemic and decision objects. Package C remains an optional activation laboratory.

The v0.2 audit-resolution record identifies the changes from v0.1. Integration must remain a separate scoped production run.
""",
        encoding="utf-8",
    )


def recursive_json_version_update(root: Path) -> None:
    def update(value: Any) -> Any:
        if isinstance(value, dict):
            out = {key: update(item) for key, item in value.items()}
            if "programme_version" in out:
                out["programme_version"] = PROGRAMME_VERSION
            if "data_origin" in out:
                out["data_origin"] = DATA_ORIGIN
            if "not_for_empirical_validation" in out:
                out["not_for_empirical_validation"] = True
            return out
        if isinstance(value, list):
            return [update(item) for item in value]
        if value == "OG-SYNTH-0.1":
            return PROGRAMME_VERSION
        return value

    for path in sorted(root.rglob("*.json")):
        try:
            value = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        write_json(update(value), path)


def create_schema_catalog(root: Path) -> None:
    rows: list[dict[str, Any]] = []
    skip = {"MANIFEST.csv", "PACKAGE_MANIFEST.csv", "data_dictionary.csv"}
    for path in sorted(root.rglob("*.csv")):
        if path.name in skip:
            continue
        try:
            df = pd.read_csv(path, nrows=50)
        except Exception:
            continue
        rel = str(path.relative_to(root))
        for column in df.columns:
            lower = column.lower()
            if column == "record_hash":
                role = "integrity"
            elif column in PROVENANCE_COLUMNS:
                role = "provenance"
            elif lower.endswith("_id") or column in {
                "project_id",
                "period_id",
                "reference_project_id",
            }:
                role = "identifier"
            elif any(
                token in lower
                for token in [
                    "p50",
                    "p80",
                    "probability",
                    "score",
                    "ratio",
                    "value",
                    "cost",
                    "duration",
                    "days",
                    "risk",
                    "cpi",
                    "spi",
                    "float",
                ]
            ):
                role = "analytical"
            else:
                role = "attribute"
            if "usd" in lower or "cost" in lower:
                unit = "USD_OR_AS_NAMED"
            elif "day" in lower or "date" in lower:
                unit = "DAY_OR_DATE"
            elif any(
                token in lower
                for token in [
                    "fraction",
                    "probability",
                    "reliability",
                    "membership",
                    "ratio",
                ]
            ):
                unit = "0_TO_1_OR_AS_NAMED"
            else:
                unit = "AS_NAMED"
            rows.append(
                {
                    "file": rel,
                    "field": column,
                    "pandas_dtype": str(df[column].dtype),
                    "role": role,
                    "description": "Synthetic programme field; see package README and module map.",
                    "unit": unit,
                }
            )
    dictionary = pd.DataFrame(rows)
    dictionary.to_csv(root / "data_dictionary.csv", index=False)
    catalog = {
        "programme_version": PROGRAMME_VERSION,
        "files": {
            filename: group[
                ["field", "pandas_dtype", "role", "unit"]
            ].to_dict("records")
            for filename, group in dictionary.groupby("file")
        },
    }
    write_json(catalog, root / "schemas" / "schema_catalog.json")


def create_package_local_manifests(root: Path) -> None:
    package_dirs = [
        root / "package_A_project_structures",
        root / "package_B_reference_training_decisions",
        root / "package_C_optional_activation_lab",
    ]
    for package_dir in package_dirs:
        for old in ["PACKAGE_MANIFEST.csv", "PACKAGE_CHECKSUMS.sha256"]:
            (package_dir / old).unlink(missing_ok=True)
        rows: list[dict[str, Any]] = []
        for path in sorted(package_dir.rglob("*")):
            if not path.is_file():
                continue
            row_count = None
            if path.suffix.lower() == ".csv":
                try:
                    row_count = len(pd.read_csv(path))
                except Exception:
                    row_count = None
            rows.append(
                {
                    "file": str(path.relative_to(package_dir)),
                    "bytes": path.stat().st_size,
                    "sha256": sha256(path),
                    "row_count": row_count,
                }
            )
        pd.DataFrame(rows).to_csv(package_dir / "PACKAGE_MANIFEST.csv", index=False)
        checksum_lines: list[str] = []
        for path in sorted(package_dir.rglob("*")):
            if not path.is_file() or path.name == "PACKAGE_CHECKSUMS.sha256":
                continue
            checksum_lines.append(
                f"{sha256(path)}  {path.relative_to(package_dir)}"
            )
        (package_dir / "PACKAGE_CHECKSUMS.sha256").write_text(
            "\n".join(checksum_lines) + "\n", encoding="utf-8"
        )


def create_root_manifest(root: Path) -> pd.DataFrame:
    (root / "MANIFEST.csv").unlink(missing_ok=True)
    (root / "CHECKSUMS.sha256").unlink(missing_ok=True)
    rows: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        row_count = None
        if path.suffix.lower() == ".csv":
            try:
                row_count = len(pd.read_csv(path))
            except Exception:
                row_count = None
        rows.append(
            {
                "file": str(path.relative_to(root)),
                "bytes": path.stat().st_size,
                "sha256": sha256(path),
                "row_count": row_count,
                "data_origin": DATA_ORIGIN,
            }
        )
    manifest = pd.DataFrame(rows)
    manifest.to_csv(root / "MANIFEST.csv", index=False)
    checksum_lines: list[str] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.name == "CHECKSUMS.sha256":
            continue
        checksum_lines.append(f"{sha256(path)}  {path.relative_to(root)}")
    (root / "CHECKSUMS.sha256").write_text(
        "\n".join(checksum_lines) + "\n", encoding="utf-8"
    )
    return manifest


def normalize_zip_file(path: Path) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        temporary, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for name in sorted(source.namelist()):
            data = source.read(name)
            if name == "docProps/core.xml":
                text = data.decode("utf-8")
                import re
                text = re.sub(
                    r"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                    '<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:type="dcterms:W3CDTF">2026-08-11T00:00:00Z</dcterms:modified>',
                    text,
                )
                data = text.encode("utf-8")
            info = zipfile.ZipInfo(name)
            info.date_time = (2026, 8, 11, 0, 0, 0)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            target.writestr(info, data)
    temporary.replace(path)


def create_workbook(root: Path, manifest: pd.DataFrame, validation: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    dark = "17365D"
    white = "FFFFFF"
    teal = "DDEBF7"
    orange = "FCE4D6"

    def format_sheet(sheet, widths: list[int] | None = None) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if widths:
            for index, width in enumerate(widths, 1):
                sheet.column_dimensions[get_column_letter(index)].width = width
        else:
            for column in range(1, sheet.max_column + 1):
                maximum = max(
                    len(str(sheet.cell(row, column).value or ""))
                    for row in range(1, min(sheet.max_row, 250) + 1)
                )
                sheet.column_dimensions[get_column_letter(column)].width = min(
                    max(maximum + 2, 12), 55
                )

    overview = workbook.create_sheet("Overview")
    overview.append(["Item", "Value"])
    for row in [
        ["Programme", "Opus Gubernatio Synthetic Evidence and Decision Programme"],
        ["Version", PROGRAMME_VERSION],
        ["Random seed", SEED],
        ["Data origin", DATA_ORIGIN],
        ["Package A", "Project-specific canonical structures including NCR and environmental cohorts"],
        ["Package B", "Reference/training, epistemic, and decision/optimization objects"],
        ["Package C", "Optional/disabled activation laboratory"],
        ["Validation", f"{'PASS' if validation['passed'] else 'FAIL'} — {validation['check_count']} checks"],
        ["Warning", "Synthetic fixtures verify implementation only; they do not establish field calibration or empirical validity."],
    ]:
        overview.append(row)
    format_sheet(overview, [28, 105])
    for row in range(2, overview.max_row + 1):
        overview.cell(row, 1).font = Font(bold=True, color="666666")
        overview.cell(row, 2).fill = PatternFill(
            "solid",
            fgColor=orange if overview.cell(row, 1).value == "Warning" else teal,
        )

    corrections = workbook.create_sheet("Audit Corrections")
    corrections.append(["Finding", "v0.2 resolution"])
    for row in [
        ["Combined archive root artefacts", "Builder, validator, report, module map, schema catalog, manifest and checksums included."],
        ["NCR Rate corpus absent", "Quality audits, NCR events and ground truth added."],
        ["Environmental Compliance corpus absent", "Requirement register, assessments, violations and ground truth added."],
        ["CCPM chain traceability and flat sizing", "Explicit chains/activity links and RSS PERT variance sizing added."],
        ["ABM agent rules absent", "Machine-readable rule table and execution trace added."],
        ["LP models prose only", "Numerical solver-consumable coefficients and bounds added."],
        ["Module numbering mismatch", "Literature-to-code alias map added."],
        ["DSM package boundary", "DSM explicitly retained in Package A as project-specific structure."],
    ]:
        corrections.append(row)
    format_sheet(corrections, [42, 105])

    module_map = pd.read_csv(root / "module_asset_map.csv")
    module_sheet = workbook.create_sheet("Module Asset Map")
    module_sheet.append(list(module_map.columns))
    for row in module_map.itertuples(index=False):
        module_sheet.append(list(row))
    format_sheet(module_sheet)

    aliases = pd.read_csv(root / "module_id_aliases.csv")
    alias_sheet = workbook.create_sheet("Module ID Aliases")
    alias_sheet.append(list(aliases.columns))
    for row in aliases.itertuples(index=False):
        alias_sheet.append(list(row))
    format_sheet(alias_sheet)

    validation_sheet = workbook.create_sheet("Validation")
    validation_sheet.append(["Check", "Result", "Detail"])
    for check in validation["checks"]:
        validation_sheet.append(
            [check["check"], "PASS" if check["passed"] else "FAIL", check["detail"]]
        )
    format_sheet(validation_sheet, [70, 12, 80])

    file_sheet = workbook.create_sheet("File Manifest")
    file_sheet.append(list(manifest.columns))
    for row in manifest.itertuples(index=False):
        file_sheet.append(list(row))
    format_sheet(file_sheet, [75, 14, 70, 14, 30])

    workbook.properties.created = FIXED_DT
    workbook.properties.modified = FIXED_DT
    path = root / "package_summary.xlsx"
    workbook.save(path)
    normalize_zip_file(path)


def deterministic_zip(source: Path, target: Path, arc_root: str) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    target.unlink(missing_ok=True)
    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(source.rglob("*")):
            if not path.is_file():
                continue
            info = zipfile.ZipInfo(str(Path(arc_root) / path.relative_to(source)))
            info.date_time = (2026, 8, 11, 0, 0, 0)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            archive.writestr(info, path.read_bytes())


def create_requirements_lock(root: Path) -> None:
    versions = []
    for module_name, distribution in [
        ("numpy", "numpy"),
        ("pandas", "pandas"),
        ("scipy", "scipy"),
        ("networkx", "networkx"),
        ("openpyxl", "openpyxl"),
    ]:
        module = __import__(module_name)
        versions.append(f"{distribution}=={getattr(module, '__version__', 'UNKNOWN')}")
    (root / "requirements-lock.txt").write_text(
        "\n".join(versions) + "\n", encoding="utf-8"
    )


def copy_tooling(root: Path, base_zip: Path) -> None:
    generators = root / "generators"
    base_dir = generators / "base"
    base_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(Path(__file__), generators / "build_opus_synthetic_programme_v0_2.py")
    script_dir = Path(__file__).resolve().parent
    shutil.copy2(
        script_dir / "validate_synthetic_programme_v0_2.py",
        generators / "validate_synthetic_programme_v0_2.py",
    )
    shutil.copy2(
        script_dir / "verify_synthetic_checksums_v0_2.py",
        generators / "verify_synthetic_checksums_v0_2.py",
    )
    shutil.copy2(base_zip, base_dir / base_zip.name)
    write_json(
        {
            "programme_version": PROGRAMME_VERSION,
            "generated_at": GENERATED_AT,
            "base_archive": base_zip.name,
            "base_archive_sha256": sha256(base_zip),
            "builder": "build_opus_synthetic_programme_v0_2.py",
            "builder_sha256": sha256(Path(__file__)),
            "validator": "validate_synthetic_programme_v0_2.py",
            "validator_sha256": sha256(script_dir / "validate_synthetic_programme_v0_2.py"),
            "random_seed": SEED,
        },
        root / "BUILD_PROVENANCE.json",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--base-zip",
        type=Path,
        default=Path("/mnt/data/Opus_Gubernatio_Synthetic_Programme_v0.1.zip"),
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("/mnt/data/Opus_Gubernatio_Synthetic_Programme_v0.2"),
    )
    parser.add_argument(
        "--combined-zip",
        type=Path,
        default=Path("/mnt/data/Opus_Gubernatio_Synthetic_Programme_v0.2.zip"),
    )
    parser.add_argument(
        "--separate-dir", type=Path, default=Path("/mnt/data")
    )
    args = parser.parse_args()

    if not args.base_zip.exists():
        raise FileNotFoundError(args.base_zip)
    if args.output_root.exists():
        shutil.rmtree(args.output_root)
    args.output_root.mkdir(parents=True)

    with tempfile.TemporaryDirectory() as temporary_directory:
        base_root = safe_extract(args.base_zip, Path(temporary_directory))
        base_map = pd.read_csv(base_root / "module_asset_map.csv", dtype={"module_id": str})
        for directory in [
            "package_A_project_structures",
            "package_B_reference_training_decisions",
            "package_C_optional_activation_lab",
        ]:
            shutil.copytree(base_root / directory, args.output_root / directory)

    # Remove any old local manifests/checksums, then refresh versioned content.
    for path in args.output_root.rglob("*"):
        if path.name in {"PACKAGE_MANIFEST.csv", "PACKAGE_CHECKSUMS.sha256"}:
            path.unlink(missing_ok=True)
    refresh_csv_provenance(args.output_root)
    refresh_json_versions(args.output_root)

    build_ccpm(args.output_root)
    build_abm_rules(args.output_root)
    build_ncr_corpus(args.output_root)
    build_environmental_corpus(args.output_root)
    build_lp_models(args.output_root)
    build_module_maps(args.output_root, base_map)
    write_readmes(args.output_root)
    recursive_json_version_update(args.output_root)
    refresh_csv_provenance(args.output_root)
    create_requirements_lock(args.output_root)
    copy_tooling(args.output_root, args.base_zip)
    create_schema_catalog(args.output_root)
    create_package_local_manifests(args.output_root)

    validator = args.output_root / "generators" / "validate_synthetic_programme_v0_2.py"
    subprocess.run(
        [
            sys.executable,
            str(validator),
            "--root",
            str(args.output_root),
            "--write-report",
        ],
        check=True,
    )
    validation = json.loads(
        (args.output_root / "validation_report.json").read_text(encoding="utf-8")
    )

    # Create manifest for workbook, then regenerate after workbook is included.
    manifest = create_root_manifest(args.output_root)
    create_workbook(args.output_root, manifest, validation)
    create_package_local_manifests(args.output_root)
    manifest = create_root_manifest(args.output_root)

    deterministic_zip(args.output_root, args.combined_zip, "Opus_Gubernatio_Synthetic_Programme_v0.2")
    deterministic_zip(
        args.output_root / "package_A_project_structures",
        args.separate_dir / "Opus_Gubernatio_Package_A_Project_Structures_v0.2.zip",
        "package_A_project_structures",
    )
    deterministic_zip(
        args.output_root / "package_B_reference_training_decisions",
        args.separate_dir
        / "Opus_Gubernatio_Package_B_Reference_Training_Decisions_v0.2.zip",
        "package_B_reference_training_decisions",
    )
    deterministic_zip(
        args.output_root / "package_C_optional_activation_lab",
        args.separate_dir / "Opus_Gubernatio_Package_C_Optional_Activation_Lab_v0.2.zip",
        "package_C_optional_activation_lab",
    )

    verifier = args.output_root / "generators" / "verify_synthetic_checksums_v0_2.py"
    subprocess.run(
        [sys.executable, str(verifier), "--root", str(args.output_root)], check=True
    )

    print(
        json.dumps(
            {
                "root": str(args.output_root),
                "combined_zip": str(args.combined_zip),
                "validation_checks": validation["check_count"],
                "validation_failures": validation["failed_count"],
                "manifest_files": len(manifest),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
