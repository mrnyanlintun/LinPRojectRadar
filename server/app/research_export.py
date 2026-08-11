"""
De-identified export and archive chain (B6), extended to two kinds and a workbook format.

This turns collected data into an analysable dataset. Three properties matter more than
anything else here.

De-identification is an allowlist, never a denylist. Every *_COLUMNS tuple below names every
field that may leave the system for its sheet; a row is assembled by naming each field
explicitly, so a column added to a model later cannot appear in an export by default. A
denylist would invert the failure: the day someone adds an ip_hash column, every export
silently starts carrying it, and the leak is discovered after the data has been shared.

The payload is regenerated on fetch rather than stored. research_exports records the checksum,
not the bytes, so fetching re-derives the payload from the current data and compares. That is a
stronger property than reading back a blob: it detects the underlying rows changing after the
export was taken, which is exactly the drift that would silently invalidate an analysis.

Two kinds, selected by the caller, with different scopes:

  PARTICIPANT_INPUTS. Per participant, filtered to research accounts server-side
  unconditionally, with a date window over DECISION completion (final_submitted_at) — a
  decision counts as belonging to the window it was completed in. This is `build_rows`,
  unchanged in name and behaviour from before this work, extended with judgement-facing
  columns (Part 5).

  PROJECT_HEALTH. Per project, with a date window over the COMPUTATION timestamp
  (ComputedResult.computed_at), not a decision timestamp — there is no decision in this
  scope, and a reporting period is an integer, not a range a date window can bound. This
  scope is NOT filtered to research accounts: ComputedResult belongs to a project, and a
  project has no account_type of its own — an operational project's analytical results are
  exactly as reachable here as a research one's. The banner and the notice both follow from
  this: see NOTICE_OPERATIONAL below and the `research_account_filtered` response field.

Free text is included and flagged. rationale is a dependent variable, so it has to be
exported, but participants can type anything into it, including their own name or a
colleague's. The export carries an explicit review flag rather than quietly shipping text
nobody has read. The workbook's analysis_long sheet carries NONE of it, by construction: it is
built from a fixed column list that contains no free-text field, so nothing needs to be
scrubbed from it after the fact.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import pathlib
import re
import zipfile
from datetime import date, datetime, timezone
from typing import Any, Callable

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .facade import err
from .research_identity import audit, resolve_caller
from .research_models import (
    Assignment, AuditEvent, Configuration, ComputedResult, Decision, DecisionSupportPackage,
    Participant, ParticipantProfile, ResearchExport, Scenario, Transition,
)
from .models import Project

EXPORT_KINDS: tuple[str, ...] = ("participant_inputs", "project_health")
EXPORT_FORMATS: tuple[str, ...] = ("json", "csv", "xlsx")

# --------------------------------------------------------------------------- Decisions sheet
#
# Every field that may leave the system for the Decisions sheet / the legacy json+csv export,
# in export order. Adding a field here is the only way to export one. Nothing is derived from
# a model's column list.
EXPORT_COLUMNS: tuple[str, ...] = (
    # identity: the pseudonymous code and nothing else
    "pseudonymous_code",
    "order_group",
    # the instance itself, for joining across sheets (Decisions / Stimulus / analysis_long)
    "instance_id",
    # design
    "scenario_id",
    "scenario_version",
    "sequence_number",
    "period",
    "config_code",
    # preliminary judgment
    "pre_action",
    "pre_confidence",
    "pre_assessment",
    "pre_submitted_at",
    "pre_locked_at",
    # reveal
    "reveal_at",
    "package_id",
    "package_version",
    "package_hash",
    # final decision
    "final_action",
    "disposition",
    "final_confidence",
    "final_submitted_at",
    "escalation_level",
    "owner_role",
    "authority_role",
    "resource_constraint",
    # T4 structured capture (migration 0011)
    "evidence_items",
    "reason_code",
    "deadline",
    # free text, flagged for review
    "rationale",
    "residual_risk",
    # transition
    "branch_id",
    "branch_version",
    "transition_seed",
    "transition_probability",
    "next_state_id",
    "transition_displayed_at",
    # derived analysis variables
    "judgment_shift_action",
    "confidence_shift",
    "deliberation_seconds",
    "pre_assessment_seconds",
    # Part 5: fields for JUDGING a case, not fields the model consumes. The data is cleaned by
    # hand; this is what a human reviewer needs to tell a real response from a broken one.
    "time_on_instance_seconds",
    "pre_committed_before_disclosure",
    "completion_state",
    "session_break",
)

# Checked by the tests against the serialised payload. These names must never appear.
FORBIDDEN_FIELDS: tuple[str, ...] = (
    "access_token_hash", "session_ref", "ip_hash", "ip_address", "email", "participant_id",
    "consent_id", "access_token",
)

# Columns whose content is participant-authored and may contain identifiers.
# T4 added residual_risk. It is participant-authored prose in exactly the way rationale is — a
# participant describing what risk they are accepting can and will name a project, a contractor,
# or a colleague — so it carries the same review flag rather than shipping unreviewed.
# reason_code and evidence_items are NOT here: the first is a closed vocabulary and the second is
# a list of labels the interface itself generated, so neither can contain free composition.
FREE_TEXT_COLUMNS: tuple[str, ...] = ("pre_assessment", "rationale", "residual_risk")

# --------------------------------------------------------------------------- Stimulus sheet
#
# One row per instance: the researcher-authored frozen package AS DISCLOSED. Exactly the fields
# `decision-ui.js` renders to a participant on reveal (`renderPackage`) plus the identity and
# timing columns needed to join back to Decisions. Nothing here is produced by the analytical
# layer — see Module results for that — and nothing here is participant-authored, so none of it
# needs the free-text review flag.
STIMULUS_COLUMNS: tuple[str, ...] = (
    "pseudonymous_code",
    "instance_id",
    "scenario_id",
    "period",
    "package_id",
    "package_version",
    "package_hash",
    "model_version",
    "use_case",
    "output_type",
    "data_cutoff",
    # the brief: the evidentiary narrative around the recommendation
    "detected_condition",
    "alternatives",
    "uncertainty",
    "limitations",
    "applicability_boundary",
    "expiration_trigger",
    "provenance",
    # the recommendation, as disclosed
    "recommended_action",
    "frozen_at",
    "reveal_at",
)

# --------------------------------------------------------------------------- Module results sheet
#
# One row per project, period and computation. Referred to by NAME and GROUP, per
# NAMING_AUTHORITY.md rule 6 ("never a module id or number in user-facing text") — this sheet
# reaches a committee, which is exactly that surface. `computation` and `group` are the only
# identifying columns; the internal new_id ("A1.1") never appears.
MODULE_RESULT_COLUMNS: tuple[str, ...] = (
    "project",
    "period",
    "computed_at",
    "computation",
    "group",
    "activation_state",
    "signal_qualification",
    "status_color",
    "evidence_metric",
    "result_json",
)

GROUP_NAMES: dict[str, str] = {
    "A": "Project Health",
    "B": "Recommendation and Governance",
    "C": "Data and Evidence Health",
    "D": "Portfolio Level",
}

# The module name table. Read independently of server/app/simulation/ (which this task must not
# modify): this is the same source file registry.py reads (p0-baseline/module_renumbering_map.csv
# is data, not code under that directory), loaded here on its own so the export has no import
# dependency on the simulation package at all.
_MODULE_NAME_CSV = (pathlib.Path(__file__).resolve().parents[2]
                    / "p0-baseline" / "module_renumbering_map.csv")

# Run 1 remediation (remediation_programme.md, remediation_decisions_answered.md 1.3, 1.4; the
# run-1 prompt Part 4). Mirrored from server/app/simulation/registry.py's DISABLED_CONCEPT_ONLY,
# CORE_VOTING_MODULES and PROXY_QUALIFIERS -- that file is the source of truth for these three
# sets; this module mirrors rather than imports them for the same reason it mirrors module
# names above, so the export keeps no import dependency on the simulation package at all. The
# export is one of the three surfaces (export, API, methods documentation) the qualifier is
# allowed on; the participant ledger and decision card are not, and neither reads this file.
_RUN1_DISABLED: dict[str, str] = {
    "A3.8": "Parametric Cost Index",
    "B2.7": "Plithogenic Sets",
    "B2.9": "Quantum Probability",
    "B2.20": "Hypersoft Sets",
    "B4.1": "Multi-Objective Optimization",
    "B4.2": "Linear Programming",
    "B4.5": "Decision Sensitivity Matrix",
    "B4.6": "Pareto Frontier Analysis",
}

_RUN1_CORE_VOTING: frozenset[str] = frozenset({
    "A1.7", "A1.8", "A2.8", "A3.2", "A3.4", "A4.2", "A4.3",
})

_RUN1_PROXY_QUALIFIERS: dict[str, str] = {
    "A1.2": "hard-coded transformations of two-sided CUSUM on real SPI history; k, H, sigma "
            "floor and Amber band uncalibrated",
    "A1.3": "Normal-normal updating with designed constant variances, not a governed Bayesian "
            "model",
    "A1.4": "Scalar Kalman recursion with fixed Q and R, short history, no calibrated filtering "
            "claim",
    "A1.9": "an expenditure-versus-progress control ratio, not a standardised statistical test",
    "A1.10": "fixed 50 per cent shrinkage toward historical mean; coefficient not estimated",
    "A2.4": "a custom compression ratio; no network-based crashing model or calibrated bands",
    "A2.6": "a single planned versus actual snapshot, not a longitudinal S-curve analysis",
    "A2.7": "a simplified shift summary on real milestone history, bands uncalibrated",
    "A3.3": "a labour-hours ratio, not an earned-output productivity model",
    "A3.5": "a transparent ratio; validity depends on whether the indirect plan is total or "
            "period-to-date",
    "A3.7": "an analogous-cost ratio; project selection, normalisation and adaptation "
            "ungoverned",
    "A3.9": "a material-escalation ratio with no external price index, time base or geography",
    # Revised by the fifteen-defects run, mirroring registry.py: the fallbacks this label named
    # no longer exist, so the label no longer names them. See the note there.
    "A4.5": "a lost-days over available-float ratio with ungoverned bands, computed only from "
            "verified lost days and a reported float figure",
    "A4.6": "contract growth plus a raw count; no time or exposure denominator",
    "A4.7": "an ad hoc 0.3 / 0.3 / 0.4 weighted sum; weights and dependence uncalibrated",
    "A4.8": "a precomputed compliance score; provenance and construction unvalidated",
    "A5.2": "local CPI perturbation plus deviations, not calibrated multivariate sensitivity",
    "A5.3": "a ranking of four present-state deviations; no outcome-response ranges estimated",
    "B2.10": "hard-coded transformations of raw CPI, SPI and document risk",
    "B2.11": "hard-coded memberships consuming raw metrics; no calibration evidenced",
    "B2.12": "designed perturbations, not elicited or observed hesitant assessments",
    "B2.13": "membership intervals that are designed constants",
    "B2.14": "entropy over designed state probabilities; measures the lookup, not the project",
    "B2.15": "fixed mappings from raw metrics; no governed possibility distribution",
    "B2.16": "algebraically bounded but fixed memberships on raw unqualified inputs",
    "B2.17": "formula-shaped with designed memberships, no empirical or elicitation basis",
    "B3.5": "a raw modification count; not a frequency without a denominator",
    "B4.3": "an explainable four-rule checklist, not a constraint-satisfaction solver",
    "B4.4": "four deterministic EAC variants; not an action-by-scenario matrix or optimiser",
    "D1.2": "an empirical CPI and SPI percentile rank; small-n behaviour and bands unvalidated",
}


# Remediation Run 3, the flat-to-nested adapter. The fourteen modules whose declared input is a
# NESTED assembled signal package, which the normal path never constructed, so they abstained on
# every real run. Mirrored from server/app/simulation/signal_package.py's NESTED_INPUT_MODULES,
# the source of truth, for the same no-import-dependency reason as the Run 1 tables above.
_RUN3_NEWLY_WIRED: frozenset[str] = frozenset({
    "B1.1", "B1.2", "B1.3", "B1.4",
    "B2.1", "B2.2", "B2.3", "B2.4", "B2.5", "B2.6", "B2.7", "B2.8", "B2.9",
    "B3.1",
})

# Audit P0 finding 2, recorded on every row rather than only in a report. The Category 9
# eligibility gate the architecture requires is not implemented anywhere in this platform: no
# module's inputs are qualified before it reads them, and evidence combination and governance
# therefore run on raw signals. It is stated for every computation, not only for the fourteen the
# adapter reached, because it is true of every computation, and marking only the fourteen would
# imply the rest are qualified.
_SIGNAL_QUALIFICATION = ("unqualified: no eligibility gate qualifies a signal package before "
                         "evidence combination and governance read it")


def _run1_activation_state(new_id: str) -> str:
    if new_id in _RUN1_DISABLED:
        return "DISABLED_UNSAFE"
    if new_id in _RUN1_CORE_VOTING:
        return "ENABLED_QUALIFIED"
    return "ADVISORY_ONLY"


def _run1_label(new_id: str, canonical_name: str) -> str:
    """The canonical name, plus its proxy qualifier or its disabled note, in the export's one
    fixed form. Every other module's label is its unmodified canonical name."""
    if new_id in _RUN1_DISABLED:
        return (f"{canonical_name} (disabled: concept-only, no production implementation of "
                "the analytical structure its name claims. Not executed, non-voting.)")
    qualifier = _RUN1_PROXY_QUALIFIERS.get(new_id)
    if qualifier is not None:
        return f"{canonical_name} (proxy: {qualifier}. Advisory, non-voting.)"
    if new_id in _RUN3_NEWLY_WIRED:
        return (f"{canonical_name} (newly wired and unvalidated: reachable on the normal "
                "computation path only since the flat-to-nested signal adapter, and not "
                "validated against real project evidence. Advisory, non-voting.)")
    return canonical_name


_module_names_cache: dict[str, str] | None = None


def _module_names() -> dict[str, str]:
    global _module_names_cache
    if _module_names_cache is None:
        names: dict[str, str] = {}
        if _MODULE_NAME_CSV.exists():
            with _MODULE_NAME_CSV.open(encoding="utf-8-sig") as fh:
                for row in csv.DictReader(fh):
                    new_id = (row.get("new_id") or "").strip()
                    name = (row.get("module_name") or "").strip()
                    if new_id and new_id.upper() != "RETIRED":
                        names[new_id] = name
        _module_names_cache = names
    return _module_names_cache


# --------------------------------------------------------------------------- analysis_long sheet
#
# Part 4. Long format, one row per participant per instance per post_ai level (0 = preliminary,
# 1 = final). A participant with twelve instances produces twenty-four rows, ALWAYS — including
# an instance whose final decision does not exist yet, which still contributes its post_ai=0 row
# and a post_ai=1 row of nulls. Omitting the second row would be a silent filter on incomplete
# instances, which Part 5 forbids explicitly.
#
# NO FREE TEXT. Every column here is short, closed-vocabulary, or numeric — there is no
# free-composition field in this list, so nothing needs scrubbing after the fact.
#
# expert_reference_score is ALWAYS EMPTY. The expert reference standard does not exist yet
# (see REPORT for the establishment of this). The column is reserved now so that adding it
# later does not change every earlier export's shape.
LONG_COLUMNS: tuple[str, ...] = (
    "participant_id",
    "instance_id",
    "post_ai",
    "action",
    "confidence",
    "scenario",
    "project",
    "period",
    "years_experience",
    "ai_familiarity",
    "timestamp",
    "expert_reference_score",
)


def _iso(value: datetime | date | None) -> str | None:
    if value is None:
        return None
    return value.isoformat()


def _seconds_between(later: datetime | None, earlier: datetime | None) -> float | None:
    if later is None or earlier is None:
        return None
    return round((later - earlier).total_seconds(), 3)


def _parse_range(payload: dict) -> tuple[datetime | None, datetime | None, str | None]:
    def one(key: str) -> tuple[datetime | None, str | None]:
        raw = payload.get(key)
        if raw in (None, ""):
            return None, None
        try:
            parsed = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        except ValueError:
            return None, f"{key} is not an ISO 8601 timestamp"
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed, None

    start, problem = one("date_from")
    if problem:
        return None, None, problem
    end, problem = one("date_to")
    if problem:
        return None, None, problem
    if start and end and start > end:
        return None, None, "date_from is after date_to"
    return start, end, None


def _assignment_start(session: Session, participant_id: str, scenario_id: str,
                      period: str) -> datetime | None:
    """
    When the participant first saw the evidence for this period.

    Taken from the earliest evidence_viewed audit event rather than from a column, because there
    is no such column: the assignment row records allocation, not when the participant opened it.
    Using the audit trail keeps the derived variable traceable to a recorded event.
    """
    rows = session.scalars(
        select(AuditEvent).where(
            AuditEvent.event_type == "evidence_viewed",
            AuditEvent.participant_id == participant_id,
            AuditEvent.scenario_id == scenario_id,
        ).order_by(AuditEvent.server_ts)
    ).all()
    for row in rows:
        meta = row.event_metadata or {}
        if str(meta.get("period") or "P1") == period:
            return row.server_ts
    return None


def _session_break(session: Session, participant_id: str,
                   window_start: datetime | None, window_end: datetime | None) -> bool | None:
    """
    Whether the participant authenticated again DURING this instance — a signal they left and
    came back, for a human reviewer to weigh, not a performance measure.

    A HEURISTIC, stated as one: a fresh `research_login` or `sso_login` audit event strictly
    between the instance's start and its end. None (not False) when the instance has no end yet
    (the window cannot be judged), so a reviewer can tell "no break detected" from "not yet
    judgeable" — collapsing them to False would misreport an in-progress instance as clean.
    """
    if window_start is None or window_end is None:
        return None
    rows = session.scalars(
        select(func.count()).select_from(AuditEvent).where(
            AuditEvent.participant_id == participant_id,
            AuditEvent.event_type.in_(("research_login", "sso_login")),
            AuditEvent.server_ts > window_start,
            AuditEvent.server_ts < window_end,
        )
    ).first()
    return bool(rows and rows > 0)


class _Instance:
    """One eligible (participant, scenario, period) instance, joined once and reused across
    every sheet that needs it, so the Decisions, Stimulus and analysis_long sheets can never
    disagree about which instances exist."""

    __slots__ = ("decision", "assignment", "participant", "scenario", "config", "package",
                "transition", "profile", "started")


def _eligible_instances(session: Session, start: datetime | None,
                        end: datetime | None) -> list[_Instance]:
    """
    Every instance for a RESEARCH account, in the date window over decision completion.

    B8 account separation, UNCONDITIONAL: only research accounts enter this scope. This is not
    a parameter, cannot be overridden by any payload field, and applies to every export ever
    taken, including refetches of exports created before this filter existed. An operational
    account's rows never leave the system through this path.
    """
    query = select(Decision).order_by(Decision.decision_id)
    if start is not None:
        query = query.where(Decision.final_submitted_at >= start)
    if end is not None:
        query = query.where(Decision.final_submitted_at <= end)

    out: list[_Instance] = []
    for decision in session.scalars(query).all():
        assignment = session.get(Assignment, decision.assignment_id)
        if assignment is None:
            continue
        participant = session.get(Participant, assignment.participant_id)
        if participant is None or participant.account_type != "research":
            continue
        inst = _Instance()
        inst.decision = decision
        inst.assignment = assignment
        inst.participant = participant
        inst.scenario = session.get(Scenario, assignment.scenario_id)
        inst.config = (session.get(Configuration, assignment.config_id)
                      if assignment.config_id else None)
        inst.package = (session.get(DecisionSupportPackage, decision.package_id)
                        if decision.package_id else None)
        inst.transition = session.scalar(
            select(Transition).where(Transition.decision_id == decision.decision_id))
        inst.profile = session.scalar(
            select(ParticipantProfile)
            .where(ParticipantProfile.participant_id == participant.participant_id)
            .order_by(ParticipantProfile.captured_at.desc())
        )
        inst.started = _assignment_start(session, assignment.participant_id,
                                         assignment.scenario_id, decision.period or "P1")
        out.append(inst)
    return out


def build_rows(session: Session, start: datetime | None, end: datetime | None) -> list[dict[str, Any]]:
    """
    One row per decision: participant x scenario x period. The Decisions sheet, and the
    entirety of the legacy json/csv participant-inputs export.

    The range filters on final_submitted_at, so a decision counts as belonging to the window in
    which it was completed. Filtering on pre_submitted_at would split a decision across windows
    when a participant paused between periods.
    """
    rows: list[dict[str, Any]] = []
    for inst in _eligible_instances(session, start, end):
        decision, assignment, participant = inst.decision, inst.assignment, inst.participant
        scenario, config, package = inst.scenario, inst.config, inst.package
        transition = inst.transition

        shift = None
        if decision.final_action is not None and decision.pre_action is not None:
            shift = decision.final_action != decision.pre_action

        confidence_shift = None
        if decision.final_confidence is not None and decision.pre_confidence is not None:
            confidence_shift = decision.final_confidence - decision.pre_confidence

        pre_committed_before_disclosure = None
        if decision.pre_locked_at is not None and decision.reveal_at is not None:
            pre_committed_before_disclosure = decision.pre_locked_at <= decision.reveal_at

        from .research_decision import derive_stage
        completion_state = derive_stage(decision)

        window_end = decision.final_submitted_at or decision.reveal_at or decision.pre_submitted_at
        session_break = _session_break(session, assignment.participant_id, inst.started,
                                       decision.final_submitted_at)

        # Assembled by naming every field. No model introspection, no dict(row), no **kwargs:
        # each of those would let a new column travel outwards without anyone deciding it should.
        row = {
            "pseudonymous_code": participant.pseudonymous_code if participant else None,
            "order_group": participant.order_group if participant else None,
            "instance_id": decision.decision_id,
            "scenario_id": assignment.scenario_id,
            "scenario_version": scenario.scenario_version if scenario else None,
            "sequence_number": assignment.sequence_number,
            "period": decision.period,
            # The analyst's view, unlike the participant's: the condition must be present.
            "config_code": config.code if config else None,
            "pre_action": decision.pre_action,
            "pre_confidence": decision.pre_confidence,
            "pre_assessment": decision.pre_assessment,
            "pre_submitted_at": _iso(decision.pre_submitted_at),
            "pre_locked_at": _iso(decision.pre_locked_at),
            "reveal_at": _iso(decision.reveal_at),
            "package_id": decision.package_id,
            "package_version": package.version if package else None,
            "package_hash": decision.package_hash,
            "final_action": decision.final_action,
            "disposition": decision.disposition,
            "final_confidence": decision.final_confidence,
            "final_submitted_at": _iso(decision.final_submitted_at),
            "escalation_level": decision.escalation_level,
            "owner_role": decision.owner_role,
            "authority_role": decision.authority_role,
            "resource_constraint": decision.resource_constraint,
            "evidence_items": decision.evidence_items,
            "reason_code": decision.reason_code,
            "deadline": decision.deadline,
            "rationale": decision.rationale,
            "residual_risk": decision.residual_risk,
            "branch_id": transition.branch_id if transition else None,
            "branch_version": transition.branch_version if transition else None,
            "transition_seed": transition.seed if transition else None,
            "transition_probability": transition.probability if transition else None,
            "next_state_id": transition.next_state_id if transition else None,
            "transition_displayed_at": _iso(transition.displayed_at) if transition else None,
            "judgment_shift_action": shift,
            "confidence_shift": confidence_shift,
            "deliberation_seconds": _seconds_between(decision.final_submitted_at,
                                                     decision.reveal_at),
            "pre_assessment_seconds": _seconds_between(decision.pre_submitted_at, inst.started),
            "time_on_instance_seconds": _seconds_between(window_end, inst.started),
            "pre_committed_before_disclosure": pre_committed_before_disclosure,
            "completion_state": completion_state,
            "session_break": session_break,
        }

        # Defensive, and cheap: a row must contain exactly the allowlist. If these ever disagree
        # the export fails rather than shipping an unexpected shape.
        if set(row) != set(EXPORT_COLUMNS):
            raise RuntimeError(
                "export row does not match EXPORT_COLUMNS; "
                f"unexpected={sorted(set(row) - set(EXPORT_COLUMNS))} "
                f"missing={sorted(set(EXPORT_COLUMNS) - set(row))}"
            )
        rows.append({k: row[k] for k in EXPORT_COLUMNS})

    return rows


def build_stimulus_rows(session: Session, start: datetime | None,
                        end: datetime | None) -> list[dict[str, Any]]:
    """
    One row per instance: the frozen package as it was actually disclosed. Same instance set
    build_rows uses (same date window, same research-account filter), so the two sheets can
    never describe a different population of instances. A row exists even when no package has
    been revealed yet — every field but the identity columns is then None, which is the honest
    "nothing shown yet" state, not an omitted row.
    """
    rows: list[dict[str, Any]] = []
    for inst in _eligible_instances(session, start, end):
        decision, assignment, participant, package = (
            inst.decision, inst.assignment, inst.participant, inst.package)
        row = {
            "pseudonymous_code": participant.pseudonymous_code if participant else None,
            "instance_id": decision.decision_id,
            "scenario_id": assignment.scenario_id,
            "period": decision.period,
            "package_id": decision.package_id,
            "package_version": package.version if package else None,
            "package_hash": decision.package_hash,
            "model_version": package.model_version if package else None,
            "use_case": package.use_case if package else None,
            "output_type": package.output_type if package else None,
            "data_cutoff": _iso(package.data_cutoff) if package else None,
            "detected_condition": package.detected_condition if package else None,
            "alternatives": package.alternatives if package else None,
            "uncertainty": package.uncertainty if package else None,
            "limitations": package.limitations if package else None,
            "applicability_boundary": package.applicability_boundary if package else None,
            "expiration_trigger": package.expiration_trigger if package else None,
            "provenance": package.provenance if package else None,
            "recommended_action": package.recommended_action if package else None,
            "frozen_at": _iso(package.frozen_at) if package else None,
            "reveal_at": _iso(decision.reveal_at),
        }
        rows.append({k: row[k] for k in STIMULUS_COLUMNS})
    return rows


def build_analysis_long_rows(session: Session, start: datetime | None,
                             end: datetime | None) -> list[dict[str, Any]]:
    """
    Part 4. Long format for statistical software: exactly two rows per instance, post_ai 0 and
    1, always — the second row is present even when the final decision does not exist yet.
    """
    rows: list[dict[str, Any]] = []
    for inst in _eligible_instances(session, start, end):
        decision, assignment, participant = inst.decision, inst.assignment, inst.participant
        scenario, profile = inst.scenario, inst.profile
        project = scenario.evidence_package_id if scenario else None
        years_experience = profile.years_experience if profile else None
        ai_familiarity_raw = profile.ai_familiarity if profile else None
        try:
            ai_familiarity: float | None = float(ai_familiarity_raw) if ai_familiarity_raw not in (
                None, "") else None
        except (TypeError, ValueError):
            ai_familiarity = None

        base = {
            "participant_id": participant.pseudonymous_code if participant else None,
            "instance_id": decision.decision_id,
            "scenario": assignment.scenario_id,
            "project": project,
            "period": decision.period,
            "years_experience": years_experience,
            "ai_familiarity": ai_familiarity,
            "expert_reference_score": None,  # reserved; see module docstring
        }
        rows.append({**base, "post_ai": 0, "action": decision.pre_action,
                    "confidence": decision.pre_confidence,
                    "timestamp": _iso(decision.pre_submitted_at)})
        rows.append({**base, "post_ai": 1, "action": decision.final_action,
                    "confidence": decision.final_confidence,
                    "timestamp": _iso(decision.final_submitted_at)})
    return [{k: r[k] for k in LONG_COLUMNS} for r in rows]


def build_module_results_rows(session: Session, project_legacy_ids: set[str] | None,
                              start: datetime | None,
                              end: datetime | None) -> list[dict[str, Any]]:
    """
    One row per project, period and computation, referred to by name and group.

    `project_legacy_ids=None` means every project (project_health scope). A restricted set
    (participant_inputs scope) is the projects the eligible instances' scenarios point at,
    via `scenario.evidence_package_id` — the analytical record BEHIND what those participants
    were shown, kept alongside their decisions.

    A TRAINING PROJECT'S RESULTS NEVER LEAVE HERE, UNCONDITIONALLY, THE SAME WAY
    `_eligible_instances` filters to research accounts. project_health has no account_type to
    filter on (see the module docstring), so a training project's ComputedResult rows are
    otherwise exactly as reachable as a real operational project's — this is the one place that
    closes it. `project_legacy_ids` restricts participant_inputs to specific evidence projects,
    which can never be training (an evidence project is named by a Scenario and training mode
    does not use scenarios), so the skip below is a no-op there and load-bearing only for
    project_health.

    The window is over `computed_at`, a real timestamp — never a decision timestamp, because
    there is no decision in this scope, and a reporting period is an integer a date range
    cannot bound (see the module docstring and the report's Part 1 discussion).

    LIVE RESULTS ONLY (`superseded_by IS NULL`): a superseded result is not the project's
    current account of that period, and duplicating both would double the projects a
    reporting period appears under with no way to tell which is current.
    """
    query = select(ComputedResult).where(ComputedResult.superseded_by.is_(None))
    if start is not None:
        query = query.where(ComputedResult.computed_at >= start)
    if end is not None:
        query = query.where(ComputedResult.computed_at <= end)
    query = query.order_by(ComputedResult.project_id, ComputedResult.period)

    names = _module_names()
    rows: list[dict[str, Any]] = []
    for result in session.scalars(query).all():
        project = session.get(Project, result.project_id)
        if project is not None and project.is_training:
            continue
        legacy = project.legacy_id if project else None
        if project_legacy_ids is not None and legacy not in project_legacy_ids:
            continue
        for module in (result.module_results or []):
            if not isinstance(module, dict):
                continue
            module_id = str(module.get("module_id") or "")
            group_letter = str(module.get("group") or "")
            extra = {k: v for k, v in module.items()
                    if k not in ("module_id", "group", "status_color", "evidence_metric")}
            canonical_name = names.get(module_id, module_id)
            rows.append({
                "project": legacy,
                "period": result.period,
                "computed_at": _iso(result.computed_at),
                "computation": _run1_label(module_id, canonical_name),
                "group": GROUP_NAMES.get(group_letter, group_letter),
                "activation_state": _run1_activation_state(module_id),
                "signal_qualification": _SIGNAL_QUALIFICATION,
                "status_color": module.get("status_color"),
                "evidence_metric": module.get("evidence_metric"),
                "result_json": json.dumps(extra, sort_keys=True, default=str),
            })
    return rows


def _project_ids_for_instances(session: Session, start: datetime | None,
                               end: datetime | None) -> set[str]:
    ids: set[str] = set()
    for inst in _eligible_instances(session, start, end):
        if inst.scenario and inst.scenario.evidence_package_id:
            ids.add(inst.scenario.evidence_package_id)
    return ids


# --------------------------------------------------------------------------- the notice
#
# AN EXPORT IS THE ARTIFACT MOST LIKELY TO BE READ WITHOUT ANY SURROUNDING CONTEXT. It leaves the
# platform as a file and reaches people who never saw the sign-in notice or the site footer.
#
# TWO KINDS, TWO NOTICE VARIANTS — established, not assumed. participant_inputs is filtered to
# research accounts unconditionally (build_rows / _eligible_instances), so the research variant's
# claims are true of everything in it: synthetic project data, a research participant's session.
# project_health is NOT filtered by account type at all — ComputedResult belongs to a project,
# and a project carries no account_type of its own, so an operational project's real, non-
# synthetic analytical results are exactly as reachable there as a research project's. Using the
# research variant's "All project data is synthetic" for project_health would be a claim this
# scope cannot back. The operational variant makes no such claim, so it is the one used there —
# selecting between the two ALREADY-approved variants by content scope, the same way every other
# surface in this codebase switches on account type; nothing here composes a third.
#
# Quoted verbatim from DISCLAIMERS_DRAFT.md. Do not edit here, do not shorten for a narrower
# format, and do not compose a variant: a surface carries the approved text whole or does not
# carry it. test_disclaimers.py fails if these diverge from the source by a character.
NOTICE_RESEARCH: tuple[str, ...] = (
    "Notice: academic research instrument. Opus Gubernatio is a proof of concept developed "
    "solely for doctoral research and demonstration. It is not a commercial service and is "
    "provided as is, without warranty of any kind, express or implied.",

    "All project data is synthetic. No real project, agency, employer, contractor, or vendor is "
    "referenced. Do not upload confidential, proprietary, personally identifiable, or otherwise "
    "sensitive information, or any document relating to an actual project.",

    "Uploaded content is sent to third-party artificial intelligence services for extraction and "
    "is stored in research infrastructure. Analytical outputs are advisory. They are not a "
    "validated compliance determination, a contractual direction, or a diagnosis of a live "
    "project. The operator disclaims all liability arising from or relating to uploaded content "
    "to the fullest extent permitted by law.",
)

NOTICE_OPERATIONAL: tuple[str, ...] = (
    "Notice. Opus Gubernatio is provided as is, without warranty of any kind, express or "
    "implied.",

    "Analytical outputs are advisory. They are not a validated compliance determination, a "
    "contractual direction, or a diagnosis of a live project.",

    "Uploaded content is sent to third-party artificial intelligence services for extraction and "
    "is stored in the platform. You are responsible for confirming that you are authorized to "
    "upload each document, and for your organization's data handling, confidentiality, and "
    "records obligations. The operator disclaims all liability arising from or relating to "
    "uploaded content to the fullest extent permitted by law.",
)

ATTRIBUTION = (
    "Developed as part of doctoral research at the School of Engineering and Applied Science, "
    "The George Washington University. The university is not a party to this notice and does not "
    "endorse or warrant the platform."
)

COPYRIGHT = (
    "© 2026 Nyan Lin Tun. All rights reserved. Opus Gubernatio and the associated software "
    "and documentation are the intellectual property of the author. Unauthorized reproduction, "
    "distribution, or use is prohibited."
)


def _notice_for(kind: str) -> tuple[str, ...]:
    return NOTICE_RESEARCH if kind == "participant_inputs" else NOTICE_OPERATIONAL


def research_account_filtered(kind: str) -> bool:
    """Whether this export's scope is unconditionally filtered to research accounts. Read by
    the frontend to decide whether the "filtered to research accounts" banner is true."""
    return kind == "participant_inputs"


def date_window_field(kind: str) -> str:
    """Which timestamp the date window bounds, for this kind. Surfaced so the UI can say so
    rather than leaving the user to guess what "From"/"To" mean for project health."""
    return "final_submitted_at" if kind == "participant_inputs" else "computed_at"


def serialise(rows: list[dict[str, Any]], fmt: str, columns: tuple[str, ...] = EXPORT_COLUMNS,
              *, include_notice: bool = True, kind: str = "participant_inputs") -> tuple[bytes, str | None]:
    """
    Render a single-table payload (json or csv). The checksum covers exactly these bytes.

    `include_notice=False` reproduces the pre-notice bytes and exists only so
    a_adminexportfetch can recognise an export taken before the notice was added. See the
    comment there. `columns` lets project_health's flat single-table json/csv export reuse this
    function over MODULE_RESULT_COLUMNS instead of EXPORT_COLUMNS.
    """
    if fmt == "json":
        body: dict[str, Any] = {
            "kind": kind,
            "columns": list(columns),
            "row_count": len(rows),
            "rows": rows,
        }
        if columns == EXPORT_COLUMNS:
            body["free_text_columns"] = list(FREE_TEXT_COLUMNS)
            body["review_required"] = bool(FREE_TEXT_COLUMNS)
            body["review_note"] = ("Free-text columns are participant-authored and may contain "
                                   "identifying content. Review before sharing outside the "
                                   "study team.")
        if include_notice:
            body["notice"] = list(_notice_for(kind))
            body["attribution"] = ATTRIBUTION
            body["copyright"] = COPYRIGHT
        return json.dumps(body, sort_keys=True, separators=(",", ":"),
                          default=str).encode("utf-8"), None

    if fmt == "csv":
        # THE CSV CARRIES NO NOTICE, AND THAT IS REPORTED RATHER THAN WORKED AROUND.
        #
        # RFC 4180 has no comment syntax. Anything placed above the header row is read as the
        # header. Repeating six hundred characters of prose in an extra column on every row is
        # not a notice, and shortening it to fit a cell is composing a new liability variant,
        # which a session may not do. So the format genuinely cannot carry the approved text.
        buffer = io.StringIO(newline="")
        writer = csv.DictWriter(buffer, fieldnames=list(columns),
                                lineterminator="\n", extrasaction="raise")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: ("" if row[k] is None else row[k]) for k in columns})
        return buffer.getvalue().encode("utf-8"), None

    return b"", f"unsupported format: {fmt}"


# --------------------------------------------------------------------------- the workbook

_FIXED_XLSX_DATE = datetime(2026, 1, 1, tzinfo=timezone.utc)


def _normalize_xlsx_bytes(raw: bytes) -> bytes:
    """
    Make openpyxl's output byte-deterministic, so the same data always produces the same
    checksum — ESTABLISHED, not assumed: measured directly (see the report) that openpyxl
    stamps `docProps/core.xml`'s created/modified timestamps at the wall clock, and that the
    zip container stamps each entry's own timestamp at the wall clock too. Two workbooks built
    a second apart from identical data therefore differ byte-for-byte before this function runs.

    Fixed here rather than by setting `workbook.properties.created/modified` alone: that neutralises
    only the docProps timestamps, not the per-entry zip timestamps, which still vary. This
    rewrites the archive with every entry's timestamp pinned to one fixed value and the
    docProps timestamps textually pinned to the same value, and re-orders entries by name so
    write order cannot introduce a difference either.
    """
    src = zipfile.ZipFile(io.BytesIO(raw))
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as out:
        for name in sorted(src.namelist()):
            data = src.read(name)
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"<dcterms:created[^>]*>[^<]*</dcterms:created>",
                    b'<dcterms:created xsi:type="dcterms:W3CDTF">2026-01-01T00:00:00Z'
                    b"</dcterms:created>", data)
                data = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2026-01-01T00:00:00Z'
                    b"</dcterms:modified>", data)
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            out.writestr(info, data)
    return out_buf.getvalue()


def _sheet_rows(columns: tuple[str, ...], rows: list[dict[str, Any]]) -> list[list[Any]]:
    out = [list(columns)]
    for row in rows:
        line = []
        for c in columns:
            v = row.get(c)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, sort_keys=True, default=str)
            line.append(v)
        out.append(line)
    return out


def build_workbook(kind: str, session: Session, start: datetime | None, end: datetime | None,
                   *, include_notice: bool = True) -> bytes:
    """
    One workbook, not separate files, so every level stays physically together under one
    checksum. Sheets are named explicitly (never relying on position — most tools default to
    the first sheet).

    participant_inputs: Notice, Decisions, Stimulus, Module results (scoped to the projects the
    eligible instances' scenarios point at), analysis_long.

    project_health: Notice, Module results (every project, scoped only by the date window over
    computed_at). No Decisions/Stimulus/analysis_long: there is no participant dimension in this
    scope, and inventing one would misrepresent what this kind reports.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    wb.properties.creator = "Opus Gubernatio"
    wb.properties.created = _FIXED_XLSX_DATE
    wb.properties.modified = _FIXED_XLSX_DATE
    wb.remove(wb.active)

    if include_notice:
        ws = wb.create_sheet("Notice")
        ws.append(["OPUS GUBERNATIO", "NOTICE"])
        ws.append([])
        for para in _notice_for(kind):
            ws.append([para])
            ws.append([])
        ws.append([ATTRIBUTION])
        ws.append([])
        ws.append([COPYRIGHT])
        ws.column_dimensions["A"].width = 118

    if kind == "participant_inputs":
        decisions = build_rows(session, start, end)
        stimulus = build_stimulus_rows(session, start, end)
        long_rows = build_analysis_long_rows(session, start, end)
        project_ids = _project_ids_for_instances(session, start, end)
        modules = build_module_results_rows(session, project_ids, start, end)

        ws = wb.create_sheet("Decisions")
        for line in _sheet_rows(EXPORT_COLUMNS, decisions):
            ws.append(line)

        ws = wb.create_sheet("Stimulus")
        for line in _sheet_rows(STIMULUS_COLUMNS, stimulus):
            ws.append(line)

        ws = wb.create_sheet("Module results")
        for line in _sheet_rows(MODULE_RESULT_COLUMNS, modules):
            ws.append(line)

        ws = wb.create_sheet("analysis_long")
        for line in _sheet_rows(LONG_COLUMNS, long_rows):
            ws.append(line)
    else:
        modules = build_module_results_rows(session, None, start, end)
        ws = wb.create_sheet("Module results")
        for line in _sheet_rows(MODULE_RESULT_COLUMNS, modules):
            ws.append(line)

    buf = io.BytesIO()
    wb.save(buf)
    return _normalize_xlsx_bytes(buf.getvalue())


def checksum(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _require_admin(session: Session, payload: dict, secret: str, action: str):
    caller, problem = resolve_caller(session, payload, secret)
    if problem:
        return None, problem
    if not caller.is_admin:
        # Rule 5: refused and audited. An attempt to export is exactly the event a research audit
        # trail should hold, whether or not it succeeded.
        audit(session, "export_action_denied", participant_id=caller.participant_id,
              action=action, role=caller.role)
        session.commit()
        return None, err("not authorized: ResearchAdmin role required")
    return caller, None


def _row_count_for(session: Session, kind: str, fmt: str, start: datetime | None,
                   end: datetime | None) -> int:
    """The count reported to the caller: the primary sheet/table's row count for this kind."""
    if kind == "participant_inputs":
        return len(build_rows(session, start, end))
    return len(build_module_results_rows(session, None, start, end))


def _build_payload(session: Session, kind: str, fmt: str, start: datetime | None,
                   end: datetime | None, *, include_notice: bool = True) -> tuple[bytes, str | None]:
    if fmt == "xlsx":
        try:
            return build_workbook(kind, session, start, end, include_notice=include_notice), None
        except RuntimeError as exc:
            return b"", str(exc)
    # json/csv: a single flat table. participant_inputs uses the Decisions shape (unchanged
    # from before this work); project_health uses the Module results shape, since it has no
    # decision-shaped data at all.
    if kind == "participant_inputs":
        rows = build_rows(session, start, end)
        return serialise(rows, fmt, EXPORT_COLUMNS, include_notice=include_notice, kind=kind)
    rows = build_module_results_rows(session, None, start, end)
    return serialise(rows, fmt, MODULE_RESULT_COLUMNS, include_notice=include_notice, kind=kind)


def a_adminexportcreate(session: Session, payload: dict, secret: str, ttl: int) -> dict[str, Any]:
    caller, problem = _require_admin(session, payload, secret, "adminexportcreate")
    if problem:
        return problem

    kind = str(payload.get("kind") or "participant_inputs").strip().lower()
    if kind not in EXPORT_KINDS:
        return err(f"kind must be one of {', '.join(EXPORT_KINDS)}")

    fmt = str(payload.get("format") or "json").strip().lower()
    if fmt not in EXPORT_FORMATS:
        return err(f"format must be one of {', '.join(EXPORT_FORMATS)}")

    start, end, problem_text = _parse_range(payload)
    if problem_text:
        return err(problem_text)

    try:
        body, problem_text = _build_payload(session, kind, fmt, start, end)
    except RuntimeError as exc:
        return err(str(exc))
    if problem_text:
        return err(problem_text)

    row_count = _row_count_for(session, kind, fmt, start, end)
    digest = checksum(body)
    date_range = f"{start.isoformat() if start else 'open'}/{end.isoformat() if end else 'open'}"

    row = ResearchExport(
        format=fmt,
        kind=kind,
        row_count=row_count,
        checksum=digest,
        destination=str(payload.get("destination") or "inline"),
        date_range=date_range,
        initiated_by=caller.participant_id,
        completed_at=func.now(),
    )
    # An export is an administrative act gated by role, not by participant consent. The consent
    # gate honours this flag rather than resolving initiated_by to a participant who, being an
    # administrator, will never have consented.
    row._admin_authorised = True
    session.add(row)
    audit(session, "export_created", participant_id=caller.participant_id,
          export_format=fmt, export_kind=kind, row_count=row_count, checksum=digest,
          date_range=date_range)
    session.commit()

    session.refresh(row)
    return {
        "ok": True,
        "export_id": row.export_id,
        "kind": kind,
        "format": fmt,
        "row_count": row_count,
        "checksum": digest,
        "date_range": date_range,
        "date_window_field": date_window_field(kind),
        "research_account_filtered": research_account_filtered(kind),
        "destination": str(payload.get("destination") or "inline"),
        "completed_at": _iso(row.completed_at),
        "review_required": bool(FREE_TEXT_COLUMNS) if kind == "participant_inputs" else False,
        "free_text_columns": list(FREE_TEXT_COLUMNS) if kind == "participant_inputs" else [],
        "columns": list(EXPORT_COLUMNS) if kind == "participant_inputs"
                  else list(MODULE_RESULT_COLUMNS),
    }


def a_adminexportlist(session: Session, payload: dict, secret: str, ttl: int) -> dict[str, Any]:
    caller, problem = _require_admin(session, payload, secret, "adminexportlist")
    if problem:
        return problem
    rows = session.scalars(select(ResearchExport).order_by(ResearchExport.export_id)).all()
    return {"ok": True, "exports": [
        {"export_id": r.export_id, "kind": r.kind or "participant_inputs", "format": r.format,
         "row_count": r.row_count, "checksum": r.checksum, "destination": r.destination,
         "date_range": r.date_range, "completed_at": _iso(r.completed_at)}
        for r in rows
    ]}


def a_adminexportfetch(session: Session, payload: dict, secret: str, ttl: int) -> dict[str, Any]:
    """
    Re-derive the payload and verify it against the stored checksum.

    A mismatch is reported loudly and the payload is withheld. It means the underlying rows have
    changed since the export was taken, so any analysis quoting that checksum no longer describes
    the data now in the database, and returning it silently would hide exactly that.
    """
    caller, problem = _require_admin(session, payload, secret, "adminexportfetch")
    if problem:
        return problem

    export_id = str(payload.get("export_id") or "").strip()
    if not export_id:
        return err("export_id is required")
    record = session.get(ResearchExport, export_id)
    if record is None:
        return err(f"export not found: {export_id}")

    kind = record.kind if record.kind in EXPORT_KINDS else "participant_inputs"

    start, end = None, None
    if record.date_range and "/" in record.date_range:
        left, _, right = record.date_range.partition("/")
        if left != "open":
            start = datetime.fromisoformat(left)
        if right != "open":
            end = datetime.fromisoformat(right)

    try:
        body, problem_text = _build_payload(session, kind, record.format or "json", start, end)
    except RuntimeError as exc:
        return err(str(exc))
    if problem_text:
        return err(problem_text)

    digest = checksum(body)
    # AN EXPORT TAKEN BEFORE THE NOTICE EXISTED IS NOT A TAMPERED EXPORT.
    #
    # The stored checksum covers the bytes the payload builder produced at the time. Adding the
    # notice (or, for xlsx, changing the sheet set) changed those bytes, so every record created
    # earlier would now fail this comparison and be withheld with a message saying the underlying
    # data had changed. That message would be false: the data is what it always was, and the
    # accusation is the opposite of the integrity guarantee this check exists to provide.
    #
    # So a mismatch is checked a second time against the pre-notice serialisation. If THAT
    # matches, the rows are provably unchanged and the record simply predates the notice. The
    # payload served is the current one, notice included, because a file leaving the platform
    # should carry it; both digests are returned so the caller can see exactly what happened.
    # A record that matches neither is a real mismatch and is still refused and audited.
    legacy = False
    if digest != record.checksum:
        legacy_body, _ = _build_payload(session, kind, record.format or "json", start, end,
                                        include_notice=False)
        if checksum(legacy_body) == record.checksum:
            legacy = True
        else:
            audit(session, "export_checksum_mismatch", participant_id=caller.participant_id,
                  export_id=export_id, stored_checksum=record.checksum, recomputed=digest)
            session.commit()
            return err(
                f"checksum verification failed for export {export_id}: stored "
                f"{record.checksum}, recomputed {digest}. The underlying data has changed since "
                f"this export was taken; the payload is withheld."
            )

    row_count = _row_count_for(session, kind, record.format or "json", start, end)
    audit(session, "export_fetched", participant_id=caller.participant_id,
          export_id=export_id, checksum=digest, row_count=row_count,
          predates_notice=legacy)
    session.commit()

    notice_in_payload = (record.format or "json") != "csv"
    result: dict[str, Any] = {
        "ok": True,
        "export_id": export_id,
        "kind": kind,
        "format": record.format,
        "row_count": row_count,
        "checksum": digest,
        "checksum_verified": True,
        # True when the record was taken before the notice was added: the rows verified against
        # the stored checksum, and the payload below carries the notice the original did not.
        "predates_notice": legacy,
        "stored_checksum": record.checksum,
        # The CSV format carries no notice; see serialise(). Stated on every fetch so it is
        # visible at the point the file is taken rather than discovered later.
        "notice_in_payload": notice_in_payload,
        "date_window_field": date_window_field(kind),
        "research_account_filtered": research_account_filtered(kind),
        "review_required": bool(FREE_TEXT_COLUMNS) if kind == "participant_inputs" else False,
        "free_text_columns": list(FREE_TEXT_COLUMNS) if kind == "participant_inputs" else [],
        "review_note": ("Free-text columns are participant-authored and may contain identifying "
                        "content. Review before sharing outside the study team."),
        "columns": list(EXPORT_COLUMNS) if kind == "participant_inputs"
                  else list(MODULE_RESULT_COLUMNS),
    }
    if record.format == "xlsx":
        import base64
        result["payload_base64"] = base64.b64encode(body).decode("ascii")
    else:
        result["payload"] = body.decode("utf-8")
    return result


EXPORT_ACTIONS: dict[str, Callable[[Session, dict, str, int], dict]] = {
    "adminexportcreate": a_adminexportcreate,
    "adminexportlist": a_adminexportlist,
    "adminexportfetch": a_adminexportfetch,
}
