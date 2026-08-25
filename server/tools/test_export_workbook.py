#!/usr/bin/env python3
"""
The export: two kinds, and the XLSX workbook (Notice / Decisions / Stimulus / Module results /
analysis_long).

Run (from server/):

    DATABASE_URL=... SESSION_SECRET=... PYTHONIOENCODING=utf-8 python tools/test_export_workbook.py

Covers, end to end through /exec plus the pure layer, and by opening the produced workbook and
reading it back:

  1. The selector produces genuinely different content for the two kinds.
  2. analysis_long: exactly two rows per instance, post_ai 0 and 1, always — including an
     instance whose final decision does not exist yet.
  3. No free text reaches analysis_long.
  4. The Notice sheet's text matches DISCLAIMERS_DRAFT.md character for character, for BOTH
     variants (research for participant_inputs, operational for project_health).
  5. A participant with no decisions produces a valid, openable, empty-but-correct file.
  6. The workbook's bytes are checksum-stable across two independent builds of identical data
     (established directly, not assumed — see the report).
  7. project_health is NOT filtered to research accounts.
  8. The checksum-legacy path (predates the current sheet set) still verifies for xlsx too.

The whole run is wrapped so a crash prints a failing RESULT line, never a clean-looking silence.
"""
from __future__ import annotations

import base64
import io
import json
import pathlib
import re
import sys
import time
from datetime import date

sys.path.insert(0, __file__.rsplit("tools", 1)[0])

RESULTS: list[tuple[bool, str, str]] = []


def check(ok: bool, label: str, detail: str = "") -> None:
    RESULTS.append((bool(ok), label, detail))
    print(f"  {'PASS' if ok else 'FAIL'}  {label}" + (f"  [{detail}]" if detail else ""))


def finish() -> None:
    failed = [r for r in RESULTS if not r[0]]
    print(f"\nRESULT: {len(RESULTS) - len(failed)}/{len(RESULTS)} checks passed")
    sys.exit(1 if failed else 0)


def main() -> None:
    import openpyxl
    from fastapi.testclient import TestClient
    from sqlalchemy import select

    import app.main as main_mod
    import app.research_export as rx
    from app.models import Project
    from app.research_identity import hash_access_token
    from app.research_models import ComputedResult, Participant

    client = TestClient(main_mod.app, raise_server_exceptions=False)
    Session = main_mod.SessionFactory

    def post(payload: dict) -> dict:
        r = client.post("/exec", content=json.dumps(payload),
                        headers={"Content-Type": "text/plain"})
        assert r.status_code == 200, f"contract violation: HTTP {r.status_code}"
        return r.json()

    ADMIN = "export-wb-admin"
    with Session() as s:
        row = s.scalar(select(Participant).where(Participant.role == "ResearchAdmin"))
        if row is None:
            s.add(Participant(pseudonymous_code="EXPWB-ADMIN", role="ResearchAdmin",
                              access_token_hash=hash_access_token(ADMIN)))
        else:
            row.access_token_hash = hash_access_token(ADMIN)
        for legacy in ("ST-WB-OPEN",):
            if s.scalar(select(Project).where(Project.legacy_id == legacy)) is None:
                s.add(Project(legacy_id=legacy, doc={"id": legacy, "name": legacy}))
        s.commit()
    admin = post({"action": "researchlogin", "access_token": ADMIN})["session_token"]

    print("\nSEED: one participant, one complete instance, one bare (consented, no decision) one")
    scenario = post({"action": "adminscenariocreate", "session_token": admin,
                     "scenario_version": "wb-v1", "period_count": 1,
                     "evidence_package_id": "ST-WB-OPEN"})["scenario_id"]
    post({"action": "adminconfigurationcreate", "session_token": admin, "code": "C0",
         "version": "v1", "freeze": True})
    post({"action": "adminsequencecreate", "session_token": admin, "order_group": "GA",
         "scenario_set": "SET-WB", "version": "v1", "positions": ["C0"], "freeze": True})
    pkg = post({"action": "adminpackagecreate", "session_token": admin, "version": "wb-pkg",
               "provider_id": "frozen-store", "model_version": "m1",
               "detected_condition": "cost overrun trending",
               "recommended_action": "Escalate", "limitations": "single scenario",
               "freeze": True})

    # Participant A: completes the instance fully.
    ca = post({"action": "adminparticipantcreate", "session_token": admin,
              "pseudonymous_code": "WB-A", "account_type": "research"})
    tok_a = post({"action": "researchlogin", "access_token": ca["access_token"]})["session_token"]
    post({"action": "consentgrant", "session_token": tok_a, "consent_version": "v1"})
    post({"action": "intakesave", "session_token": tok_a,
         "responses": {"experience_level": "mid", "years_experience": 11,
                      "ai_familiarity": "4"}})
    post({"action": "adminassign", "session_token": admin, "participant_id": ca["participant_id"],
         "order_group": "GA", "scenario_set": "SET-WB", "scenario_ids": [scenario]})
    with Session() as s:
        from app.research_models import Assignment
        aid_a = s.scalar(select(Assignment).where(
            Assignment.participant_id == ca["participant_id"])).assignment_id
    post({"action": "adminpackageattach", "session_token": admin, "assignment_id": aid_a,
         "package_id": pkg["package_id"]})
    post({"action": "researchevidenceget", "session_token": tok_a})
    time.sleep(0.05)
    post({"action": "researchprejudgment", "session_token": tok_a, "pre_action": "monitor",
         "pre_confidence": 55, "pre_assessment": "seems stable so far"})
    post({"action": "researchreveal", "session_token": tok_a})
    time.sleep(0.05)
    post({"action": "researchdecision", "session_token": tok_a, "final_action": "escalate",
         "disposition": "modify", "final_confidence": 80,
         "rationale": "the trend line worried me, name: John Contractor"})

    # Participant B: consented, intake done, ASSIGNED, but never even opened the evidence —
    # "consented and decided nothing".
    cb = post({"action": "adminparticipantcreate", "session_token": admin,
              "pseudonymous_code": "WB-B", "account_type": "research"})
    tok_b = post({"action": "researchlogin", "access_token": cb["access_token"]})["session_token"]
    post({"action": "consentgrant", "session_token": tok_b, "consent_version": "v1"})
    post({"action": "adminassign", "session_token": admin, "participant_id": cb["participant_id"],
         "order_group": "GA", "scenario_set": "SET-WB", "scenario_ids": [scenario]})
    # Participant C: submits the PRELIMINARY judgment (a Decision row now exists) and stops —
    # never revealed, never decided. This is the case that proves post_ai=1 still emits.
    cc = post({"action": "adminparticipantcreate", "session_token": admin,
              "pseudonymous_code": "WB-C", "account_type": "research"})
    tok_c = post({"action": "researchlogin", "access_token": cc["access_token"]})["session_token"]
    post({"action": "consentgrant", "session_token": tok_c, "consent_version": "v1"})
    post({"action": "intakesave", "session_token": tok_c,
         "responses": {"experience_level": "junior", "years_experience": 2}})
    post({"action": "adminassign", "session_token": admin, "participant_id": cc["participant_id"],
         "order_group": "GA", "scenario_set": "SET-WB", "scenario_ids": [scenario]})
    post({"action": "researchevidenceget", "session_token": tok_c})
    post({"action": "researchprejudgment", "session_token": tok_c, "pre_action": "monitor",
         "pre_confidence": 40})

    check(True, "fixture seeded: one complete instance, one abandoned mid-way, "
          "one with nothing decided")

    # A ComputedResult row on the SAME project, so participant_inputs' Module results sheet has
    # something to scope in, and project_health has something to see at all.
    with Session() as s:
        proj = s.scalar(select(Project).where(Project.legacy_id == "ST-WB-OPEN"))
        s.add(ComputedResult(
            project_id=proj.id, period=1,
            signal_inputs={"cpi": 0.9}, simulation_version="test-1", seed="1",
            period_cutoff=date(2026, 6, 30),
            module_results=[
                {"module_id": "A1.1", "group": "A", "status_color": "amber",
                 "evidence_metric": "CPI trending down"},
                {"module_id": "B4.7", "group": "B", "status_color": "red",
                 "evidence_metric": "regret index high"},
            ],
        ))
        # A SECOND, unrelated operational-flavoured project — nothing ties it to any research
        # participant — to prove project_health is not filtered to research accounts.
        if s.scalar(select(Project).where(Project.legacy_id == "ST-WB-OPS")) is None:
            s.add(Project(legacy_id="ST-WB-OPS", doc={"id": "ST-WB-OPS", "name": "ops only"}))
            s.flush()
        ops_proj = s.scalar(select(Project).where(Project.legacy_id == "ST-WB-OPS"))
        s.add(ComputedResult(
            project_id=ops_proj.id, period=1,
            signal_inputs={"cpi": 1.1}, simulation_version="test-1", seed="2",
            period_cutoff=date(2026, 6, 30),
            module_results=[
                {"module_id": "C1.2", "group": "C", "status_color": "green",
                 "evidence_metric": "on time"},
            ],
        ))
        s.commit()

    # ================================================================== 1. genuinely different
    print("\n1. The selector produces genuinely different content for the two kinds")
    pi = post({"action": "adminexportcreate", "session_token": admin, "format": "xlsx",
              "kind": "participant_inputs"})
    ph = post({"action": "adminexportcreate", "session_token": admin, "format": "xlsx",
              "kind": "project_health"})
    check(pi.get("ok") is True and ph.get("ok") is True, "both exports created",
          f"{pi.get('ok')} {ph.get('ok')}")
    check(pi.get("research_account_filtered") is True,
          "participant_inputs reports itself as filtered to research accounts", "")
    check(ph.get("research_account_filtered") is False,
          "project_health reports itself as NOT filtered", "")
    check(pi.get("date_window_field") == "final_submitted_at",
          "participant_inputs windows on decision completion", pi.get("date_window_field"))
    check(ph.get("date_window_field") == "computed_at",
          "project_health windows on the computation timestamp", ph.get("date_window_field"))
    check(pi.get("columns") != ph.get("columns"),
          "the two kinds declare different column sets", "")

    pi_f = post({"action": "adminexportfetch", "session_token": admin,
                "export_id": pi["export_id"]})
    ph_f = post({"action": "adminexportfetch", "session_token": admin,
                "export_id": ph["export_id"]})
    pi_bytes = base64.b64decode(pi_f["payload_base64"])
    ph_bytes = base64.b64decode(ph_f["payload_base64"])
    check(pi_bytes != ph_bytes, "the two workbooks are byte-different", "")
    pi_wb = openpyxl.load_workbook(io.BytesIO(pi_bytes))
    ph_wb = openpyxl.load_workbook(io.BytesIO(ph_bytes))
    check(set(pi_wb.sheetnames) == {"Notice", "Decisions", "Stimulus", "Module results",
                                    "analysis_long"},
          "participant_inputs has all five named sheets", str(pi_wb.sheetnames))
    check(set(ph_wb.sheetnames) == {"Notice", "Module results"},
          "project_health has only Notice and Module results — no participant sheets",
          str(ph_wb.sheetnames))
    ph_ops_rows = [r for r in ph_wb["Module results"].iter_rows(min_row=2, values_only=True)
                  if r[0] == "ST-WB-OPS"]
    check(len(ph_ops_rows) == 1,
          "project_health includes the operational-only project's result", str(len(ph_ops_rows)))
    pi_module_rows = [r for r in pi_wb["Module results"].iter_rows(min_row=2, values_only=True)]
    check(all(r[0] == "ST-WB-OPEN" for r in pi_module_rows) and len(pi_module_rows) >= 1,
          "participant_inputs' Module results is scoped to the project its instances touched",
          str(pi_module_rows))

    # ================================================================== 2/3. analysis_long
    print("\n2. analysis_long: two rows per instance, post_ai 0 and 1, always")
    long_rows = list(pi_wb["analysis_long"].iter_rows(min_row=2, values_only=True))
    long_header = [c.value for c in pi_wb["analysis_long"][1]]
    check(long_header == list(rx.LONG_COLUMNS), "the header matches LONG_COLUMNS exactly",
          str(long_header))
    idx = {name: i for i, name in enumerate(long_header)}
    by_instance: dict[str, list] = {}
    for r in long_rows:
        by_instance.setdefault(r[idx["instance_id"]], []).append(r)
    check(all(len(v) == 2 for v in by_instance.values()),
          "every instance contributes exactly two rows", str({k: len(v) for k, v in
                                                              by_instance.items()}))
    post_ai_sets = {tuple(sorted(r[idx["post_ai"]] for r in v)) for v in by_instance.values()}
    check(post_ai_sets == {(0, 1)}, "post_ai is 0 and 1 for every instance", str(post_ai_sets))
    # Participant B (no decision at all) contributes NO instance — an instance is anchored on a
    # Decision row, which only exists once a preliminary judgment is submitted.
    check("WB-B" not in [r[idx["participant_id"]] for r in long_rows],
          "a participant who decided nothing contributes no analysis_long rows", "")
    a_rows = [r for r in long_rows if r[idx["participant_id"]] == "WB-A"]
    check(len(a_rows) == 2, "participant A's single instance produced exactly two rows",
          str(len(a_rows)))
    c_rows = [r for r in long_rows if r[idx["participant_id"]] == "WB-C"]
    check(len(c_rows) == 2,
          "participant C's ABANDONED instance (preliminary only) STILL produces two rows",
          str(len(c_rows)))
    c_post = [r for r in c_rows if r[idx["post_ai"]] == 1]
    check(len(c_post) == 1 and c_post[0][idx["action"]] is None
          and c_post[0][idx["confidence"]] is None,
          "its post_ai=1 row exists with null action/confidence, not omitted",
          str(c_post[0]) if c_post else "missing")
    pre_row = [r for r in a_rows if r[idx["post_ai"]] == 0][0]
    post_row = [r for r in a_rows if r[idx["post_ai"]] == 1][0]
    check(pre_row[idx["action"]] == "monitor" and pre_row[idx["confidence"]] == 55,
          "post_ai=0 carries the preliminary action and confidence",
          f"{pre_row[idx['action']]} {pre_row[idx['confidence']]}")
    check(post_row[idx["action"]] == "escalate" and post_row[idx["confidence"]] == 80,
          "post_ai=1 carries the final action and confidence",
          f"{post_row[idx['action']]} {post_row[idx['confidence']]}")
    check(pre_row[idx["years_experience"]] == 11, "years_experience carried onto the long row",
          str(pre_row[idx["years_experience"]]))
    check(pre_row[idx["ai_familiarity"]] == 4.0, "ai_familiarity carried onto the long row",
          str(pre_row[idx["ai_familiarity"]]))
    check(pre_row[idx["project"]] == "ST-WB-OPEN", "project resolved from the scenario",
          str(pre_row[idx["project"]]))
    check(pre_row[idx["expert_reference_score"]] is None,
          "expert_reference_score is present as a column and empty", "")

    print("\n3. No free text reaches analysis_long")
    forbidden_strings = ("John Contractor", "seems stable so far",
                         "the trend line worried me")
    long_blob = json.dumps(long_rows, default=str)
    for needle in forbidden_strings:
        check(needle not in long_blob, f"'{needle}' does not appear in analysis_long", "")
    check(not (set(long_header) & set(rx.FREE_TEXT_COLUMNS)),
          "LONG_COLUMNS shares no name with a free-text column", str(long_header))

    # ================================================================== 4. notice text
    # RUN 59, PHASE B. RE-POINTED AT A NON-MARKDOWN ORACLE.
    #
    # Owner's ruling, 2026-08-25: no markdown document in this repository carries authority. This
    # block compared the exported workbook's Notice sheet with blockquotes parsed out of
    # DISCLAIMERS_DRAFT.md. The real subject is production: the XLSX is a file that LEAVES the
    # platform and is read by people who never saw a footer, and what matters is that it carries
    # the SAME characters the browser surfaces carry. `assets/js/disclaimers.js` is where those
    # characters live in production -- it is the shipped constant the upload panels render from,
    # and export.js writes the Notice sheet from `window.LinDisclaimers`. So the oracle is now
    # that file. Same assertion, production source, still fails on one character.
    #
    # DISCLAIMERS_DRAFT.md is not deleted, not moved and not edited. It simply no longer decides
    # whether a check is red.
    print("\n4. The Notice sheet matches assets/js/disclaimers.js character for character")
    _js = (pathlib.Path(__file__).resolve().parents[2] / "assets" / "js"
           / "disclaimers.js").read_text(encoding="utf-8")

    def _js_paras(var: str) -> list[str]:
        import ast as _ast
        m = re.search(r"var\s+%s\s*=\s*\[(.*?)\];" % var, _js, re.S)
        assert m is not None, "disclaimers.js no longer declares var %s" % var
        lits = re.findall(r'"((?:[^"\\]|\\.)*)"', m.group(1))
        return [re.sub(r"\s+", " ", _ast.literal_eval('"%s"' % s)).strip() for s in lits]

    def _retired_run59_blockquote(n: int) -> list[str]:
        """RETIRED BY RUN 59, NOT DELETED. The markdown parse is kept verbatim below so the
        retirement is reversible and the record of what it did survives. It is not called."""
        draft = pathlib.Path(__file__).resolve().parents[2] / "DISCLAIMERS_DRAFT.md"
        draft_text = draft.read_text(encoding="utf-8")
        section = draft_text.split(f"## {n}.")[1].split("\n## ")[0]
        paras, current = [], []
        for line in section.splitlines():
            s = line.strip()
            if s.startswith(">"):
                s = s[1:].strip()
                if s == "" or s.startswith("**") and current:
                    if current:
                        paras.append(" ".join(current)); current = []
                    if s and not s.startswith("**"):
                        current.append(s)
                else:
                    current.append(s)
            elif current:
                paras.append(" ".join(current)); current = []
        if current:
            paras.append(" ".join(current))
        cleaned = [re.sub(r"\s+", " ", p).strip() for p in paras if p.strip()]
        # Markdown bold on the lead sentence ("**Notice: ...**") is source formatting, not text
        # the sheet carries — the sheet holds the plain sentence. Strip it before comparing.
        return [re.sub(r"\*\*(.+?)\*\*", r"\1", p) for p in cleaned]

    research_source = _js_paras("RESEARCH")
    operational_source = _js_paras("OPERATIONAL")
    check(len(research_source) == 3 and len(operational_source) == 3,
          "disclaimers.js still ships three research and three operational paragraphs, so the "
          "comparisons below cannot be vacuously satisfied by an empty list",
          f"{len(research_source)} / {len(operational_source)}")
    notice_pi_ws = pi_wb["Notice"]
    notice_pi_text = " ".join(str(c[0].value) for c in notice_pi_ws.iter_rows() if c[0].value)
    for para in research_source:
        norm = re.sub(r"\s+", " ", para).strip()
        check(norm in re.sub(r"\s+", " ", notice_pi_text),
              "participant_inputs Notice sheet carries a research paragraph verbatim",
              norm[:50])
    notice_ph_ws = ph_wb["Notice"]
    notice_ph_text = " ".join(str(c[0].value) for c in notice_ph_ws.iter_rows() if c[0].value)
    for para in operational_source:
        norm = re.sub(r"\s+", " ", para).strip()
        check(norm in re.sub(r"\s+", " ", notice_ph_text),
              "project_health Notice sheet carries an operational paragraph verbatim",
              norm[:50])
    check(pi_wb.sheetnames[0] == "Notice" and ph_wb.sheetnames[0] == "Notice",
          "Notice is the FIRST sheet in both workbooks", "")

    # ================================================================== 5. empty but valid
    print("\n5. A participant with no decisions in range still produces a valid file")
    empty = post({"action": "adminexportcreate", "session_token": admin, "format": "xlsx",
                 "kind": "participant_inputs",
                 "date_from": "1990-01-01T00:00:00Z", "date_to": "1990-01-02T00:00:00Z"})
    check(empty.get("ok") is True, "an out-of-range window still produces a file, not an error",
          str(empty)[:120])
    check(empty.get("row_count") == 0, "row_count is 0", str(empty.get("row_count")))
    empty_f = post({"action": "adminexportfetch", "session_token": admin,
                   "export_id": empty["export_id"]})
    empty_bytes = base64.b64decode(empty_f["payload_base64"])
    empty_wb = openpyxl.load_workbook(io.BytesIO(empty_bytes))
    check(set(empty_wb.sheetnames) == {"Notice", "Decisions", "Stimulus", "Module results",
                                       "analysis_long"},
          "the empty export still has all five sheets, headers included", str(empty_wb.sheetnames))
    check([c.value for c in empty_wb["Decisions"][1]] == list(rx.EXPORT_COLUMNS),
          "the Decisions sheet still has its header row with zero data rows",
          str(empty_wb["Decisions"].max_row))
    check(empty_wb["Decisions"].max_row == 1,
          "and genuinely zero data rows, not a blank placeholder row",
          str(empty_wb["Decisions"].max_row))

    # ================================================================== 6. determinism
    print("\n6. The workbook's bytes are checksum-stable across independent builds")
    with Session() as s:
        b1 = rx.build_workbook("participant_inputs", s, None, None)
    time.sleep(1.2)
    with Session() as s:
        b2 = rx.build_workbook("participant_inputs", s, None, None)
    check(b1 == b2, "two builds of identical data are byte-identical",
          f"{rx.checksum(b1)[:16]} vs {rx.checksum(b2)[:16]}")

    # ================================================================== 7. constraints
    print("\n7. Import constraints on analysis_long")
    check(all(" " not in c for c in rx.LONG_COLUMNS),
          "every long-sheet column name is space-free", str(rx.LONG_COLUMNS))
    check(pi_wb["analysis_long"]["A1"].value is not None,
          "no blank leading row before the header", "")
    check(len(pi_wb["Decisions"].merged_cells.ranges) == 0 and
          len(pi_wb["analysis_long"].merged_cells.ranges) == 0,
          "no merged cells on either sheet", "")

    # ================================================================== 8. legacy checksum path
    print("\n8. An xlsx export taken before this sheet set still verifies")
    from sqlalchemy import text as sqltext
    with Session() as s:
        legacy_body, _ = rx._build_payload(s, "participant_inputs", "xlsx", None, None,
                                           include_notice=False)
    with Session() as s:
        s.execute(sqltext("UPDATE research_exports SET checksum = :c WHERE export_id = :e"),
                 {"c": rx.checksum(legacy_body), "e": pi["export_id"]})
        s.commit()
    old_f = post({"action": "adminexportfetch", "session_token": admin,
                 "export_id": pi["export_id"]})
    check(old_f.get("ok") is True, "the legacy-checksum record still fetches", str(old_f)[:100])
    check(old_f.get("predates_notice") is True, "and is reported as predating the notice", "")


try:
    main()
except Exception as e:  # a crash must read as a FAILURE, never as a clean run
    import traceback
    traceback.print_exc()
    check(False, f"suite crashed: {type(e).__name__}: {e}")
finish()
